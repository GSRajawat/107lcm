import streamlit as st
import pandas as pd
import datetime
import os
import io
import zipfile
import tempfile
import fitz  # PyMuPDF
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import json
import ast
from sqlalchemy import create_engine

# --- Load PostgreSQL URL from Streamlit secrets ---
try:
    pg_url = st.secrets["connections"]["postgres"]["url"]
    engine = create_engine(pg_url)
except KeyError:
    st.error("❌ Database URL not found. Please set it in Streamlit secrets.")
    st.stop()

# --- Helper function to upload CSV to Supabase ---
def upload_csv_to_supabase(csv_path, table_name):
    if not os.path.exists(csv_path):
        st.warning(f"⚠️ File not found: {csv_path}")
        return
    try:
        df = pd.read_csv(csv_path)
        if df.empty:
            st.warning(f"⚠️ File exists but is empty: {csv_path}")
            return
        st.write(f"📄 Preview of `{table_name}` data:", df.head())
        st.info(f"Uploading {len(df)} rows to table `{table_name}`...")
        df.to_sql(table_name, engine, if_exists="replace", index=False, method='multi', chunksize=100)
        st.success(f"✅ Uploaded `{table_name}` to Supabase PostgreSQL!")
    except Exception as e:
        st.error(f"❌ Failed to upload `{table_name}`: {e}")


# --- Buttons to trigger upload ---
if st.button("⬆️ Upload Timetable to Supabase"):
    upload_csv_to_supabase("timetable.csv", "timetable")

if st.button("⬆️ Upload Attestation Data to Supabase"):
    upload_csv_to_supabase("attestation_data_combined.csv", "attestation_data")


# --- Configuration ---
CS_REPORTS_FILE = "cs_reports.csv"
EXAM_TEAM_MEMBERS_FILE = "exam_team_members.csv"
SHIFT_ASSIGNMENTS_FILE = "shift_assignments.csv"
ROOM_INVIGILATORS_FILE = "room_invigilator_assignments.csv" # New file for room-wise invigilators
SITTING_PLAN_FILE = "sitting_plan.csv" # Standardized sitting plan filename
TIMETABLE_FILE = "timetable.csv" # Standardized timetable filename
ASSIGNED_SEATS_FILE = "assigned_seats.csv" # New file for assigned seats
ATTESTATION_DATA_FILE = "attestation_data_combined.csv" # For rasa_pdf output
COLLEGE_STATISTICS_FILE = "college_statistics_fancy.csv" # For college_statistic output

# Helper function to ensure consistent string formatting for paper codes (remove .0 if numeric)
def _format_paper_code(code_str):
    if pd.isna(code_str) or not code_str:
        return ""
    s = str(code_str).strip()
    # If it looks like a float (e.g., "12345.0"), convert to int string
    if s.endswith('.0') and s[:-2].isdigit():
        return s[:-2]
    return s

def load_shift_assignments():
    if os.path.exists(SHIFT_ASSIGNMENTS_FILE):
        try:
            df = pd.read_csv(SHIFT_ASSIGNMENTS_FILE)
            # Convert string representations of lists back to actual lists for roles
            for role in ["senior_center_superintendent", "center_superintendent", "assistant_center_superintendent", 
                         "permanent_invigilator", "assistant_permanent_invigilator", 
                         "class_3_worker", "class_4_worker"]: # Added new roles
                if role in df.columns:
                    # Ensure that empty strings or NaN values are handled gracefully
                    df[role] = df[role].apply(lambda x: ast.literal_eval(x) if pd.notna(x) and x.strip() else [])
            return df
        except Exception as e:
            st.error(f"Error loading shift assignments: {e}. Reinitializing shift assignments file.")
            # If an error occurs during loading, reinitialize the DataFrame with correct columns
            return pd.DataFrame(columns=['date', 'shift', 'senior_center_superintendent', 'center_superintendent', 
                                         "assistant_center_superintendent", "permanent_invigilator", 
                                         "assistant_permanent_invigilator", "class_3_worker", "class_4_worker"]) # Added new columns here
    # If file does not exist, create a new DataFrame with all columns
    return pd.DataFrame(columns=['date', 'shift', 'senior_center_superintendent', 'center_superintendent', 
                                 "assistant_center_superintendent", "permanent_invigilator", 
                                 "assistant_permanent_invigilator", "class_3_worker", "class_4_worker"]) # Added new columns here

def save_shift_assignment(date, shift, assignments):
    assignments_df = load_shift_assignments()
    
    # Create a unique key for the assignment
    assignment_key = f"{date}_{shift}"

    # Prepare data for DataFrame, converting lists to string representations
    data_for_df = {
        'date': date,
        'shift': shift,
        'senior_center_superintendent': str(assignments.get('senior_center_superintendent', [])),
        'center_superintendent': str(assignments.get('center_superintendent', [])), 
        'assistant_center_superintendent': str(assignments.get('assistant_center_superintendent', [])),
        'permanent_invigilator': str(assignments.get('permanent_invigilator', [])),
        'assistant_permanent_invigilator': str(assignments.get('assistant_permanent_invigilator', [])),
        'class_3_worker': str(assignments.get('class_3_worker', [])), # Added new role
        'class_4_worker': str(assignments.get('class_4_worker', []))  # Added new role
    }
    new_row_df = pd.DataFrame([data_for_df])

    # Check if assignment_key already exists
    if assignment_key in (assignments_df['date'] + '_' + assignments_df['shift']).values:
        # Update existing record
        idx_to_update = assignments_df[(assignments_df['date'] == date) & (assignments_df['shift'] == shift)].index[0]
        for col, val in data_for_df.items():
            assignments_df.loc[idx_to_update, col] = val
    else:
        # Add new record
        assignments_df = pd.concat([assignments_df, new_row_df], ignore_index=True)
    
    try:
        assignments_df.to_csv(SHIFT_ASSIGNMENTS_FILE, index=False)
        return True, "Shift assignments saved successfully!"
    except Exception as e:
        return False, f"Error saving shift assignments: {e}"


# Load data from CSV files (sitting_plan.csv, timetable.csv, assigned_seats.csv)
def load_data():
    # Check if files exist before attempting to read them
    sitting_plan_df = pd.DataFrame()
    timetable_df = pd.DataFrame()
    assigned_seats_df = pd.DataFrame() # Initialize assigned_seats_df

    if os.path.exists(SITTING_PLAN_FILE):
        try:
            sitting_plan_df = pd.read_csv(SITTING_PLAN_FILE, dtype={
                f"Roll Number {i}": str for i in range(1, 11)
            })
            sitting_plan_df.columns = sitting_plan_df.columns.str.strip()
            # Ensure Paper Code column is consistently formatted in sitting_plan_df
            if 'Paper Code' in sitting_plan_df.columns:
                sitting_plan_df['Paper Code'] = sitting_plan_df['Paper Code'].apply(_format_paper_code)
        except Exception as e:
            st.error(f"Error loading {SITTING_PLAN_FILE}: {e}")
            sitting_plan_df = pd.DataFrame()


    if os.path.exists(TIMETABLE_FILE):
        try:
            timetable_df = pd.read_csv(TIMETABLE_FILE)
            timetable_df.columns = timetable_df.columns.str.strip()
            # Ensure Paper Code column is consistently formatted in timetable_df
            if 'Paper Code' in timetable_df.columns:
                timetable_df['Paper Code'] = timetable_df['Paper Code'].apply(_format_paper_code)
        except Exception as e:
            st.error(f"Error loading {TIMETABLE_FILE}: {e}")
            timetable_df = pd.DataFrame()
    
    if os.path.exists(ASSIGNED_SEATS_FILE):
        try:
            # Ensure Room Number and Roll Number are read as string to prevent type mismatch issues
            # Also ensure Paper Code is read as string and formatted
            assigned_seats_df = pd.read_csv(ASSIGNED_SEATS_FILE, dtype={"Roll Number": str, "Room Number": str, "Paper Code": str, "Date": str, "Shift": str})
            if 'Paper Code' in assigned_seats_df.columns:
                assigned_seats_df['Paper Code'] = assigned_seats_df['Paper Code'].apply(_format_paper_code)
        except Exception as e:
            st.error(f"Error loading {ASSIGNED_SEATS_FILE}: {e}")
            assigned_seats_df = pd.DataFrame(columns=["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "Date", "Shift"])
    else:
        assigned_seats_df = pd.DataFrame(columns=["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "Date", "Shift"])
            
    return sitting_plan_df, timetable_df, assigned_seats_df

# Save uploaded files (for admin panel)
def save_uploaded_file(uploaded_file_content, filename):
    try:
        if isinstance(uploaded_file_content, pd.DataFrame):
            # If it's a DataFrame, convert to CSV bytes
            csv_bytes = uploaded_file_content.to_csv(index=False).encode('utf-8')
        else:
            # Assume it's bytes from st.file_uploader
            # Ensure uploaded_file_content is a BytesIO object or similar with .getbuffer()
            if hasattr(uploaded_file_content, 'getbuffer'):
                csv_bytes = uploaded_file_content.getbuffer()
            else:
                # Fallback for other file-like objects, or if it's already bytes
                csv_bytes = uploaded_file_content.read()


        with open(filename, "wb") as f:
            f.write(csv_bytes)
        return True, f"File {filename} saved successfully!" # Modified: Return a tuple here
    except Exception as e:
        return False, f"Error saving file {filename}: {e}"

# Admin login (simple hardcoded credentials)
def admin_login():
    user = st.text_input("Username", type="default")
    pwd = st.text_input("Password", type="password")
    return user == "admin" and pwd == "admin123"

# Centre Superintendent login (simple hardcoded credentials)
def cs_login():
    user = st.text_input("CS Username", type="default")
    pwd = st.text_input("CS Password", type="password")
    return user == "cs_admin" and pwd == "cs_pass123"

# --- CSV Helper Functions for CS Reports ---
def load_cs_reports_csv():
    if os.path.exists(CS_REPORTS_FILE):
        try:
            df = pd.read_csv(CS_REPORTS_FILE)
            # Ensure 'class' column exists, add if missing with empty string as default
            if 'class' not in df.columns:
                df['class'] = ""
            
            # Convert string representations of lists back to actual lists
            for col in ['absent_roll_numbers', 'ufm_roll_numbers']:
                if col in df.columns:
                    # Convert to string, then handle 'nan' and empty strings before literal_eval
                    df[col] = df[col].astype(str).apply(
                        lambda x: ast.literal_eval(x) if x.strip() and x.strip().lower() != 'nan' else []
                    )
            return df
        except Exception as e:
            st.error(f"Error loading CS reports from CSV: {e}")
            return pd.DataFrame(columns=['report_key', 'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'class', 'absent_roll_numbers', 'ufm_roll_numbers'])
    else:
        return pd.DataFrame(columns=['report_key', 'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'class', 'absent_roll_numbers', 'ufm_roll_numbers'])

def save_cs_report_csv(report_key, data):
    reports_df = load_cs_reports_csv()
    
    # Convert lists to string representation for CSV storage
    data_for_df = data.copy()
    data_for_df['absent_roll_numbers'] = str(data_for_df.get('absent_roll_numbers', []))
    data_for_df['ufm_roll_numbers'] = str(data_for_df.get('ufm_roll_numbers', []))

    # Convert the single data dictionary to a DataFrame row
    new_row_df = pd.DataFrame([data_for_df])

    # Check if report_key already exists in the DataFrame
    if report_key in reports_df['report_key'].values:
        # Update existing record
        idx_to_update = reports_df[reports_df['report_key'] == report_key].index[0]
        # Update values in that row using .loc
        for col, val in data_for_df.items():
            reports_df.loc[idx_to_update, col] = val
    else:
        # Add new record
        reports_df = pd.concat([reports_df, new_row_df], ignore_index=True)

    try:
        reports_df.to_csv(CS_REPORTS_FILE, index=False)
        return True, "Report saved to CSV successfully!"
    except Exception as e:
        return False, f"Error saving report to CSV: {e}"

def load_single_cs_report_csv(report_key):
    reports_df = load_cs_reports_csv()
    filtered_df = reports_df[reports_df['report_key'] == report_key]
    if not filtered_df.empty:
        return True, filtered_df.iloc[0].to_dict()
    else:
        return False, {}

# --- Exam Team Members Functions ---
def load_exam_team_members():
    if os.path.exists(EXAM_TEAM_MEMBERS_FILE):
        try:
            df = pd.read_csv(EXAM_TEAM_MEMBERS_FILE)
            return df['name'].tolist()
        except Exception as e:
            st.error(f"Error loading exam team members: {e}")
            return []
    return []

def save_exam_team_members(members):
    df = pd.DataFrame({'name': sorted(list(set(members)))}) # Remove duplicates and sort
    try:
        df.to_csv(EXAM_TEAM_MEMBERS_FILE, index=False)
        return True, "Exam team members saved successfully!"
    except Exception as e:
        return False, f"Error saving exam team members: {e}"

# Refactored helper function to get raw student data for a session
def _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable_df):
    """
    Collects raw student data for a given date and shift from assigned_seats_df
    and merges with timetable info.
    Returns a list of dictionaries, each representing an assigned student.
    """
    all_students_data = []

    # Filter timetable for the given date and shift
    current_day_exams_tt = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) &
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if current_day_exams_tt.empty:
        return all_students_data # Return empty list if no exams found

    # Iterate through each exam scheduled for the date/shift in the timetable
    for _, tt_row in current_day_exams_tt.iterrows():
        tt_class = str(tt_row["Class"]).strip()
        tt_paper_code = str(tt_row["Paper Code"]).strip()
        tt_paper_name = str(tt_row["Paper Name"]).strip()

        # Filter assigned_seats_df for students assigned to this specific exam session
        current_exam_assigned_students = assigned_seats_df[
            (assigned_seats_df["Date"].astype(str).str.strip() == date_str) &
            (assigned_seats_df["Shift"].astype(str).str.strip().str.lower() == shift.lower()) &
            (assigned_seats_df["Paper Code"].astype(str).str.strip() == tt_paper_code) & # Use formatted paper code
            (assigned_seats_df["Paper Name"].astype(str).str.strip() == tt_paper_name)
        ]

        for _, assigned_row in current_exam_assigned_students.iterrows():
            roll_num = str(assigned_row["Roll Number"]).strip()
            room_num = str(assigned_row["Room Number"]).strip()
            seat_num_raw = str(assigned_row["Seat Number"]).strip()

            seat_num_display = ""
            seat_num_sort_key = None
            try:
                # Handle alphanumeric seats for sorting (e.g., 1A, 2A, 1B, 2B)
                if re.match(r'^\d+[A-Z]$', seat_num_raw):
                    num_part = int(re.match(r'^(\d+)', seat_num_raw).group(1))
                    char_part = re.search(r'([A-Z])$', seat_num_raw).group(1)
                    # Assign a tuple for sorting: (char_order, number)
                    seat_num_sort_key = (ord(char_part), num_part)
                    seat_num_display = seat_num_raw
                elif seat_num_raw.isdigit():
                    seat_num_sort_key = (float('inf'), int(seat_num_raw)) # Numeric seats after alphanumeric
                    seat_num_display = str(int(float(seat_num_raw))) # Display as integer string
                else:
                    seat_num_sort_key = (float('inf'), float('inf')) # Fallback for other formats
                    seat_num_display = seat_num_raw if seat_num_raw else "N/A"
            except ValueError:
                seat_num_sort_key = (float('inf'), float('inf')) # Fallback for other formats
                seat_num_display = seat_num_raw if seat_num_raw else "N/A"

            all_students_data.append({
                "roll_num": roll_num,
                "room_num": room_num,
                "seat_num_display": seat_num_display, # This is what will be displayed/exported
                "seat_num_sort_key": seat_num_sort_key, # This is for sorting
                "paper_name": tt_paper_name,
                "paper_code": tt_paper_code,
                "class_name": tt_class,
                "date": date_str,
                "shift": shift
            })
    return all_students_data

def get_all_students_for_date_shift_formatted(date_str, shift, assigned_seats_df, timetable):
    all_students_data = _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable)

    if not all_students_data:
        return None, "No students found for the selected date and shift.", None

    # Sort the collected data by Room Number, then Seat Number
    all_students_data.sort(key=lambda x: (x['room_num'], x['seat_num_sort_key']))

    # Extract exam_time and class_summary_header from timetable (similar to original logic)
    current_day_exams_tt = timetable[
        (timetable["Date"].astype(str).str.strip() == date_str) &
        (timetable["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]
    exam_time = current_day_exams_tt.iloc[0]["Time"].strip() if "Time" in current_day_exams_tt.columns else "TBD"
    unique_classes = current_day_exams_tt['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # --- Prepare text output ---
    output_string_parts = []
    output_string_parts.append("जीवाजी विश्वविद्यालय ग्वालियर")
    output_string_parts.append("परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107")
    output_string_parts.append(class_summary_header)
    output_string_parts.append(f"दिनांक :-{date_str}")
    output_string_parts.append(f"पाली :-{shift}")
    output_string_parts.append(f"समय :-{exam_time}")

    students_by_room = {}
    for student in all_students_data:
        room = student['room_num']
        if room not in students_by_room:
            students_by_room[room] = []
        students_by_room[room].append(student)

    for room_num in sorted(students_by_room.keys()):
        output_string_parts.append(f" कक्ष :-{room_num}") # Added space for consistency
        current_room_students = students_by_room[room_num]

        num_cols = 10

        for i in range(0, len(current_room_students), num_cols):
            block_students = current_room_students[i : i + num_cols]

            # Create a single line for 10 students
            single_line_students = []
            for student in block_students:
                # Modified formatting here: removed space after '(' and added '-' before paper_name
                single_line_students.append(
                    f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']})-{student['paper_name']}"
                )

            output_string_parts.append("".join(single_line_students)) # Join directly without spaces

    final_text_output = "\n".join(output_string_parts)

    # --- Prepare Excel output data ---
    excel_output_data = []

    # Excel Header
    excel_output_data.append(["जीवाजी विश्वविद्यालय ग्वालियर"])
    excel_output_data.append(["परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107"])
    excel_output_data.append([class_summary_header])
    excel_output_data.append([]) # Blank line
    excel_output_data.append(["दिनांक :-", date_str])
    excel_output_data.append(["पाली :-", shift])
    excel_output_data.append(["समय :-", exam_time])
    excel_output_data.append([]) # Blank line

    # Excel Student Data Section (now each block of 10 students is one row, each student is one cell)
    for room_num in sorted(students_by_room.keys()):
        excel_output_data.append([f" कक्ष :-{room_num}"]) # Added space for consistency
        current_room_students = students_by_room[room_num]

        num_cols = 10

        for i in range(0, len(current_room_students), num_cols):
            block_students = current_room_students[i : i + num_cols]

            excel_row_for_students = [""] * num_cols # Prepare 10 cells for this row

            for k, student in enumerate(block_students):
                # Each cell contains the full student string, modified formatting
                excel_row_for_students[k] = (
                    f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']})-{student['paper_name']}"
                )

            excel_output_data.append(excel_row_for_students)
            excel_output_data.append([""] * num_cols) # Blank row for spacing

    return final_text_output, None, excel_output_data	

# --- Room Invigilator Assignment Functions (NEW) ---
def load_room_invigilator_assignments():
    if os.path.exists(ROOM_INVIGILATORS_FILE):
        try:
            df = pd.read_csv(ROOM_INVIGILATORS_FILE)
            if 'invigilators' in df.columns:
                df['invigilators'] = df['invigilators'].astype(str).apply(
                    lambda x: ast.literal_eval(x) if x.strip() and x.strip().lower() != 'nan' else []
                )
            return df
        except Exception as e:
            st.error(f"Error loading room invigilator assignments: {e}")
            return pd.DataFrame(columns=['date', 'shift', 'room_num', 'invigilators'])
    return pd.DataFrame(columns=['date', 'shift', 'room_num', 'invigilators'])

def save_room_invigilator_assignment(date, shift, room_num, invigilators):
    inv_df = load_room_invigilator_assignments()
    
    # Create a unique key for the assignment
    assignment_key = f"{date}_{shift}_{room_num}"

    # Prepare data for DataFrame, converting list to string representation
    data_for_df = {
        'date': date,
        'shift': shift,
        'room_num': room_num,
        'invigilators': str(invigilators)
    }
    new_row_df = pd.DataFrame([data_for_df])

    # Check if assignment_key already exists
    if assignment_key in (inv_df['date'] + '_' + inv_df['shift'] + '_' + inv_df['room_num'].astype(str)).values:
        # Update existing record
        idx_to_update = inv_df[
            (inv_df['date'] == date) & 
            (inv_df['shift'] == shift) & 
            (inv_df['room_num'].astype(str) == str(room_num))
        ].index[0]
        for col, val in data_for_df.items():
            inv_df.loc[idx_to_update, col] = val
    else:
        # Add new record
        inv_df = pd.concat([inv_df, new_row_df], ignore_index=True)
    
    try:
        inv_df.to_csv(ROOM_INVIGILATORS_FILE, index=False)
        return True, "Room invigilator assignments saved successfully!"
    except Exception as e:
        return False, f"Error saving room invigilator assignments: {e}"


# Get all exams for a roll number (Student View)
def get_all_exams(roll_number, sitting_plan, timetable):
    student_exams = []
    roll_number_str = str(roll_number).strip() # Ensure consistent string comparison

    # Iterate through each row of the sitting plan
    for _, sp_row in sitting_plan.iterrows():
        # Check all possible roll number columns in the current sitting plan row
        for i in range(1, 11):
            r_col = f"Roll Number {i}"
            if r_col in sp_row and str(sp_row[r_col]).strip() == roll_number_str:
                # If roll number matches, extract paper and class details from this sitting plan row
                paper = str(sp_row["Paper"]).strip()
                paper_code = str(sp_row["Paper Code"]).strip()
                paper_name = str(sp_row["Paper Name"]).strip()
                _class = str(sp_row["Class"]).strip()

                # Find all matching entries in the timetable for this paper and class
                matches_in_timetable = timetable[
                    (timetable["Paper"].astype(str).str.strip() == paper) &
                    (timetable["Paper Code"].astype(str).str.strip() == paper_code) &
                    (timetable["Paper Name"].astype(str).str.strip() == paper_name) &
                    (timetable["Class"].astype(str).str.strip().str.lower() == _class.lower())
                ]

                # Add all found timetable matches for this student's paper to the list
                for _, tt_row in matches_in_timetable.iterrows():
                    student_exams.append({
                        "Date": tt_row["Date"],
                        "Shift": tt_row["Shift"],
                        "Class": _class,
                        "Paper": paper,
                        "Paper Code": paper_code,
                        "Paper Name": paper_name
                    })
                # Break from inner loop once the roll number is found in a row to avoid duplicate processing for the same row
                # if the roll number appears in multiple 'Roll Number X' columns within the *same* row (unlikely but safe)
                break
    return student_exams

# Get sitting details for a specific roll number and date (Student View)
def get_sitting_details(roll_number, date, sitting_plan, timetable):
    found_sittings = []
    roll_number_str = str(roll_number).strip()
    date_str = str(date).strip() # Ensure date is in 'DD-MM-YYYY' string format

    for _, sp_row in sitting_plan.iterrows():
        for i in range(1, 11):
            r_col = f"Roll Number {i}"
            s_col = f"Seat Number {i}" # Corresponding seat number column

            # Check if the roll number exists in any of the roll number columns in this sitting plan row
            if r_col in sp_row and str(sp_row[r_col]).strip() == roll_number_str:
                # If roll number matches, extract paper and class details from this sitting plan row
                paper = str(sp_row["Paper"]).strip()
                paper_code = str(sp_row["Paper Code"]).strip()
                paper_name = str(sp_row["Paper Name"]).strip()
                _class = str(sp_row["Class"]).strip()

                # Find if this paper's date matches the search in the timetable
                matches_in_timetable = timetable[
                    (timetable["Class"].astype(str).str.strip().str.lower() == _class.lower()) &
                    (timetable["Paper"].astype(str).str.strip() == paper) &
                    (timetable["Paper Code"].astype(str).str.strip() == paper_code) &
                    (timetable["Paper Name"].astype(str).str.strip() == paper_name) &
                    (timetable["Date"].astype(str).str.strip() == date_str) # Match against the provided date
                ]

                if not matches_in_timetable.empty:
                    # If there are multiple timetable entries for the same paper/class/date (e.g., different shifts),
                    # add all of them as separate sitting details.
                    for _, tt_row in matches_in_timetable.iterrows():
                        # Safely get seat number for display and sorting
                        seat_num_display = ""
                        seat_num_sort_key = float('inf') # Default sort key for non-numeric

                        if s_col in sp_row.index: # Check if column exists
                            seat_num_raw = str(sp_row[s_col]).strip()
                            try:
                                seat_num_sort_key = int(float(seat_num_raw)) # Convert to float first to handle .0, then int
                                seat_num_display = str(int(float(seat_num_raw))) # Display as integer string
                            except ValueError:
                                seat_num_display = seat_num_raw if seat_num_raw else "N/A"
                        else:
                            seat_num_display = "N/A" # Column itself is missing

                        found_sittings.append({
                            "Room Number": sp_row["Room Number"],
                            "Seat Number": seat_num_display, # Use display value
                            "Class": _class,
                            "Paper": paper,
                            "Paper Code": paper_code,
                            "Paper Name": paper_name,
                            "Date": tt_row["Date"],
                            "Shift": tt_row["Shift"],
                            "Mode": sp_row.get("Mode", ""), # Use .get() for safe access
                            "Type": sp_row.get("Type", "") # Use .get() for safe access
                        })
    return found_sittings

# New function to get all students for a given date and shift, sorted by roll number (Admin Panel)
def get_all_students_roll_number_wise_formatted(date_str, shift, assigned_seats_df, timetable):
    all_students_data = _get_session_students_raw_data(date_str, shift, assigned_seats_df, timetable)
    
    if not all_students_data:
        return None, "No students found for the selected date and shift.", None

    # Sort the collected data by Roll Number (lexicographically as strings)
    all_students_data.sort(key=lambda x: x['roll_num'])

    # Extract exam_time and class_summary_header from timetable (similar to original logic)
    current_day_exams_tt = timetable[
        (timetable["Date"].astype(str).str.strip() == date_str) &
        (timetable["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]
    exam_time = current_day_exams_tt.iloc[0]["Time"].strip() if "Time" in current_day_exams_tt.columns else "TBD"
    unique_classes = current_day_exams_tt['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # --- Prepare text output ---
    output_string_parts = []
    output_string_parts.append("जीवाजी विश्वविद्यालय ग्वालियर")
    output_string_parts.append("परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107")
    output_string_parts.append(class_summary_header)
    output_string_parts.append(f"दिनांक :-{date_str}")
    output_string_parts.append(f"पाली :-{shift}")
    output_string_parts.append(f"समय :-{exam_time}")
    output_string_parts.append("") # Blank line for separation

    num_cols = 10 
    for i in range(0, len(all_students_data), num_cols):
        block_students = all_students_data[i : i + num_cols]
        
        single_line_students = []
        for student in block_students:
            single_line_students.append(
                f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']}){student['paper_name']}"
            )
        output_string_parts.append("".join(single_line_students))

    final_text_output = "\n".join(output_string_parts)

    # --- Prepare Excel output data ---
    excel_output_data = []

    # Excel Header
    excel_output_data.append(["जीवाजी विश्वविद्यालय ग्वालियर"])
    excel_output_data.append(["परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107"])
    excel_output_data.append([class_summary_header])
    excel_output_data.append([]) # Blank line
    excel_output_data.append(["दिनांक :-", date_str])
    excel_output_data.append(["पाली :-", shift])
    excel_output_data.append(["समय :-", exam_time])
    excel_output_data.append([]) # Blank line

    # Excel Student Data Section
    for i in range(0, len(all_students_data), num_cols):
        block_students = all_students_data[i : i + num_cols]
        
        excel_row_for_students = [""] * num_cols

        for k, student in enumerate(block_students):
            excel_row_for_students[k] = (
                f"{student['roll_num']}( कक्ष-{student['room_num']}-सीट-{student['seat_num_display']}){student['paper_name']}"
            )
        
        excel_output_data.append(excel_row_for_students)
        excel_output_data.append([""] * num_cols) # Blank row for spacing

    return final_text_output, None, excel_output_data

# New helper function based on pdftocsv.py's extract_metadata, but using "UNSPECIFIED" defaults
def extract_metadata_from_pdf_text(text): 
    # Extract Class Group and Year like "BSC", "2YEAR" 
    class_match = re.search(r'([A-Z]+)\s*/?\s*(\d+(SEM|YEAR))', text) 
    class_val = f"{class_match.group(1)} {class_match.group(2)}" if class_match else "UNSPECIFIED_CLASS" 
 
    # Use regex to extract mode and type from the structured pattern
    # Looking for pattern like "BA / 1YEAR / PRIVATE / SUPP / MAR-2025"
    pattern_match = re.search(r'([A-Z]+)\s*/\s*(\d+(?:SEM|YEAR))\s*/\s*([A-Z]+)\s*/\s*([A-Z]+)\s*/\s*MAR-2025', text)
    
    if pattern_match:
        mode_type = pattern_match.group(3)  # Third element (PRIVATE/REGULAR)
        type_type = pattern_match.group(4)  # Fourth element (SUPP/REGULAR/etc)
    else:
        # Fallback to your original logic but with better ordering
        mode_type = "UNSPECIFIED_MODE" 
        # Check for PRIVATE first since it's more specific
        for keyword_mode in ["PRIVATE", "REGULAR"]: 
            if keyword_mode in text.upper(): 
                mode_type = keyword_mode 
                break 
                
        type_type = "UNSPECIFIED_TYPE" 
        # Check for more specific types first
        for keyword_type in ["ATKT", "SUPP", "EXR", "REGULAR", "PRIVATE"]: 
            if keyword_type in text.upper(): 
                type_type = keyword_type 
                break 

    paper_code = re.search(r'Paper Code[:\s]*([A-Z0-9]+)', text, re.IGNORECASE)
    paper_code_val = _format_paper_code(paper_code.group(1)) if paper_code else "UNSPECIFIED_PAPER_CODE" # Use formatter
    
    paper_name = re.search(r'Paper Name[:\s]*(.+?)(?:\n|$)', text)
    paper_name_val = paper_name.group(1).strip() if paper_name else "UNSPECIFIED_PAPER_NAME"
    
    return { 
        "class": class_val, 
        "mode": mode_type, 
        "type": type_type,  
        "room_number": "", 
        "seat_numbers": [""] * 10, 
        "paper_code": paper_code_val, 
        "paper_name": paper_name_val 
    }
  






# --- Integration of pdftocsv.py logic ---
def process_sitting_plan_pdfs(zip_file_buffer, output_sitting_plan_path, output_timetable_path):
    all_rows = []
    sitting_plan_columns = [f"Roll Number {i+1}" for i in range(10)]
    sitting_plan_columns += ["Class", "Mode", "Type", "Room Number"]
    sitting_plan_columns += [f"Seat Number {i+1}" for i in range(10)]
    sitting_plan_columns += ["Paper", "Paper Code", "Paper Name"]

    def extract_roll_numbers(text):
        # Use a set to automatically handle duplicates during extraction
        return sorted(list(set(re.findall(r'\b\d{9}\b', text)))) # De-duplicate and sort

    def format_sitting_plan_rows(rolls, paper_folder_name, meta):
        rows = []
        for i in range(0, len(rolls), 10):
            row = rolls[i:i+10]
            while len(row) < 10:
                row.append("")  # pad to ensure 10 roll number columns
            row.extend([
                meta["class"],
                meta["mode"],
                meta["type"],
                meta["room_number"]
            ])
            row.extend(meta["seat_numbers"]) # These are initially blank, filled later by assignment
            row.append(paper_folder_name)  # Use folder name as Paper
            row.append(meta["paper_code"])
            row.append(meta["paper_name"])
            rows.append(row)
        return rows

    unique_exams_for_timetable = [] # To collect data for incomplete timetable

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_file_buffer, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)
        
        base_dir = tmpdir
        # Check if there's a 'pdf_folder' sub-directory inside the extracted content
        # This handles cases where the zip contains 'pdf_folder' directly or files/folders at root
        extracted_contents = os.listdir(tmpdir)
        if 'pdf_folder' in extracted_contents and os.path.isdir(os.path.join(tmpdir, 'pdf_folder')):
            base_dir = os.path.join(tmpdir, 'pdf_folder')
        elif len(extracted_contents) == 1 and os.path.isdir(os.path.join(tmpdir, extracted_contents[0])):
            # If there's only one folder at the root, assume it's the base_dir
            base_dir = os.path.join(tmpdir, extracted_contents[0])


        processed_files_count = 0
        for folder_name in os.listdir(base_dir):
            folder_path = os.path.join(base_dir, folder_name)
            if os.path.isdir(folder_path):
                for file in os.listdir(folder_path):
                    if file.lower().endswith(".pdf"):
                        pdf_path = os.path.join(folder_path, file)
                        try:
                            doc = fitz.open(pdf_path)
                            full_text = "\n".join(page.get_text() for page in doc)
                            doc.close()
                            
                            # Use the new extract_metadata_from_pdf_text function
                            current_meta = extract_metadata_from_pdf_text(full_text)
                            
                            # Ensure paper_code and paper_name fallback to folder_name if still unspecified
                            if current_meta['paper_code'] == "UNSPECIFIED_PAPER_CODE":
                                current_meta['paper_code'] = folder_name
                            if current_meta['paper_name'] == "UNSPECIFIED_PAPER_NAME":
                                current_meta['paper_name'] = folder_name

                            rolls = extract_roll_numbers(full_text) # This now de-duplicates and sorts
                            rows = format_sitting_plan_rows(rolls, paper_folder_name=folder_name, meta=current_meta)
                            all_rows.extend(rows)
                            processed_files_count += 1
                            st.info(f"✔ Processed: {file} ({len(rolls)} unique roll numbers)")

                            # Collect unique exam details for timetable generation
                            unique_exams_for_timetable.append({
                                'Class': current_meta['class'],
                                'Paper': folder_name, # Use folder name as Paper
                                'Paper Code': current_meta['paper_code'],
                                'Paper Name': current_meta['paper_name']
                            })

                        except Exception as e:
                            st.error(f"❌ Failed to process {file}: {e}")
    
    # --- Sitting Plan Update Logic ---
    if all_rows:
        df_new_sitting_plan = pd.DataFrame(all_rows, columns=sitting_plan_columns)

        # Load existing sitting plan data
        existing_sitting_plan_df = pd.DataFrame()
        if os.path.exists(output_sitting_plan_path):
            try:
                existing_sitting_plan_df = pd.read_csv(output_sitting_plan_path, dtype={
                    f"Roll Number {i}": str for i in range(1, 11)
                })
                existing_sitting_plan_df.columns = existing_sitting_plan_df.columns.str.strip()
                if 'Paper Code' in existing_sitting_plan_df.columns:
                    existing_sitting_plan_df['Paper Code'] = existing_sitting_plan_df['Paper Code'].apply(_format_paper_code)
            except Exception as e:
                st.warning(f"Could not load existing sitting plan data for update: {e}. Starting fresh for sitting plan.")
                existing_sitting_plan_df = pd.DataFrame(columns=sitting_plan_columns)

        # Ensure all columns are present in both DataFrames before concatenation
        # Add missing columns to df_new_sitting_plan from existing_sitting_plan_df
        for col in existing_sitting_plan_df.columns:
            if col not in df_new_sitting_plan.columns:
                df_new_sitting_plan[col] = pd.NA
        # Add missing columns to existing_sitting_plan_df from df_new_sitting_plan
        for col in df_new_sitting_plan.columns:
            if col not in existing_sitting_plan_df.columns:
                existing_sitting_plan_df[col] = pd.NA

        # Reorder columns to match existing_sitting_plan_df before concatenation
        df_new_sitting_plan = df_new_sitting_plan[existing_sitting_plan_df.columns]

        # Concatenate and remove duplicates
        combined_sitting_plan_df = pd.concat([existing_sitting_plan_df, df_new_sitting_plan], ignore_index=True)

        # Define columns for identifying unique sitting plan entries.
        roll_num_cols = [f"Roll Number {i+1}" for i in range(10)]
        
        # Using all relevant columns to define uniqueness for sitting plan entries
        subset_cols_sitting_plan = roll_num_cols + ["Class", "Mode", "Type", "Room Number", "Paper", "Paper Code", "Paper Name"]
        
        # Filter subset_cols_sitting_plan to only include columns actually present in the DataFrame
        existing_subset_cols_sitting_plan = [col for col in subset_cols_sitting_plan if col in combined_sitting_plan_df.columns]

        # Fill NaN values with empty strings before dropping duplicates for consistent hashing
        combined_sitting_plan_df_filled = combined_sitting_plan_df.fillna('')
        df_sitting_plan_final = combined_sitting_plan_df_filled.drop_duplicates(subset=existing_subset_cols_sitting_plan, keep='first')

        df_sitting_plan_final.to_csv(output_sitting_plan_path, index=False)
        st.success(f"Successfully processed {processed_files_count} PDFs and updated sitting plan to {output_sitting_plan_path}")
    else:
        st.warning("No roll numbers extracted from PDFs to update sitting plan.")

    # --- Timetable Update Logic ---
    if unique_exams_for_timetable:
        df_new_timetable_entries = pd.DataFrame(unique_exams_for_timetable).drop_duplicates().reset_index(drop=True)

        # Define expected structure
        expected_columns = ["SN", "Date", "Shift", "Time", "Class", "Paper", "Paper Code", "Paper Name"]

        # Load existing timetable if exists
        if os.path.exists(output_timetable_path):
            try:
                existing_timetable_df = pd.read_csv(output_timetable_path)
                existing_timetable_df.columns = existing_timetable_df.columns.str.strip()
                if 'Paper Code' in existing_timetable_df.columns:
                    existing_timetable_df['Paper Code'] = existing_timetable_df['Paper Code'].astype(str).str.strip()
            except Exception as e:
                st.warning(f"Could not load existing timetable: {e}. Starting fresh.")
                existing_timetable_df = pd.DataFrame(columns=expected_columns)
        else:
            existing_timetable_df = pd.DataFrame(columns=expected_columns)

        # Add missing columns to both DataFrames
        for col in expected_columns:
            if col not in df_new_timetable_entries.columns:
                df_new_timetable_entries[col] = pd.NA
            if col not in existing_timetable_df.columns:
                existing_timetable_df[col] = pd.NA

        # Reorder columns
        df_new_timetable_entries = df_new_timetable_entries[expected_columns]
        existing_timetable_df = existing_timetable_df[expected_columns]

        # Concatenate and deduplicate using relevant fields
        combined_df = pd.concat([existing_timetable_df, df_new_timetable_entries], ignore_index=True)

        # Fields that define uniqueness of a timetable entry (excluding SN)
        unique_fields = ["Date", "Shift", "Time", "Class", "Paper", "Paper Code", "Paper Name"]

        # Remove duplicates based on content
        df_timetable_final = combined_df.drop_duplicates(subset=unique_fields, keep='first').reset_index(drop=True)

        # Reassign serial numbers
        df_timetable_final["SN"] = range(1, len(df_timetable_final) + 1)

        # Save final CSV
        df_timetable_final.to_csv(output_timetable_path, index=False)
        st.success(f"Timetable updated at {output_timetable_path}.")
        return True, "Timetable deduplicated and saved successfully."
    
    else:
        st.warning("No unique exam details found to generate timetable.")
        return False, "No data to process."   
    return True, "PDF processing complete."   


# --- Integration of rasa_pdf.py logic ---
def process_attestation_pdfs(zip_file_buffer, output_csv_path):
    all_data = []

    def parse_pdf_content(text):
        students = re.split(r"\n?RollNo\.\:\s*", text)
        students = [s.strip() for s in students if s.strip()]

        student_records = []

        for s in students:
            lines = s.splitlines()
            lines = [line.strip() for line in lines if line.strip()]

            def extract_after(label):
                for i, line in enumerate(lines):
                    if line.startswith(label):
                        value = line.replace(label, "", 1).strip() # Use count=1 for replace
                        if value:
                            return value
                        elif i+1 < len(lines):
                            return lines[i+1].strip()
                    # Special handling for "Regular/Backlog" as it might be on the next line
                    if label == "Regular/ Backlog:" and line.startswith("Regular/Backlog"):
                        value = line.replace("Regular/Backlog", "", 1).strip() # Use count=1 for replace
                        if value:
                            return value
                        elif i+1 < len(lines):
                            return lines[i+1].strip()
                return "" # Return empty string if label not found or value is empty

            roll_no = re.match(r"(\d{9})", lines[0]).group(1) if lines and re.match(r"(\d{9})", lines[0]) else ""
            enrollment = extract_after("Enrollment No.:")
            session = extract_after("Session:")
            regular = extract_after("Regular/ Backlog:")
            student_name = extract_after("Name:")
            father = extract_after("Father's Name:")
            mother = extract_after("Mother's Name:")
            gender = extract_after("Gender:")
            exam_name = extract_after("Exam Name:")
            centre = extract_after("Exam Centre:")
            college = extract_after("College Nmae:") # Note: Original script had 'Nmae'
            address = extract_after("Address:")

            papers = re.findall(r"([^\n]+?\[\d{5}\][^\n]*)", s) # Corrected regex for paper code

            student_data = {
                "Roll Number": roll_no,
                "Enrollment Number": enrollment,
                "Session": session,
                "Regular/Backlog": regular,
                "Name": student_name,
                "Father's Name": father,
                "Mother's Name": mother,
                "Gender": gender,
                "Exam Name": exam_name,
                "Exam Centre": centre,
                "College Name": college,
                "Address": address
            }

            for i, paper in enumerate(papers[:10]):
                student_data[f"Paper {i+1}"] = paper.strip()

            student_records.append(student_data)
        return student_records

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(zip_file_buffer, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)
        
        # Assuming PDFs are directly in the extracted folder or a subfolder named 'rasa_pdf'
        pdf_base_dir = tmpdir
        if 'rasa_pdf' in os.listdir(tmpdir) and os.path.isdir(os.path.join(tmpdir, 'rasa_pdf')):
            pdf_base_dir = os.path.join(tmpdir, 'rasa_pdf')

        processed_files_count = 0
        for filename in os.listdir(pdf_base_dir):
            if filename.lower().endswith(".pdf"):
                pdf_path = os.path.join(pdf_base_dir, filename)
                try:
                    doc = fitz.open(pdf_path)
                    text = "\n".join([page.get_text() for page in doc])
                    doc.close()
                    st.info(f"📄 Extracting: {filename}")
                    all_data.extend(parse_pdf_content(text))
                    processed_files_count += 1
                except Exception as e:
                    st.error(f"❌ Failed to process {filename}: {e}")
    
    if all_data:
        df = pd.DataFrame(all_data)
        df.to_csv(output_csv_path, index=False)
        return True, f"Successfully processed {processed_files_count} attestation PDFs and saved to {output_csv_path}"
    else:
        return False, "No data extracted from attestation PDFs."

# --- Integration of college_statistic.py logic ---
def generate_college_statistics(input_csv_path, output_csv_path):
    if not os.path.exists(input_csv_path):
        return False, f"Input file not found: {input_csv_path}. Please process attestation PDFs first."

    try:
        # Load data
        df = pd.read_csv(input_csv_path, dtype={"Roll Number": str, "Enrollment Number": str})

        # Basic cleaning
        df['College Name'] = df['College Name'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
        df['Exam Name'] = df['Exam Name'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
        df['Regular/Backlog'] = df['Regular/Backlog'].astype(str).str.strip().str.upper()

        # Extract class group and year
        def extract_class_group_and_year(exam_name):
            if pd.isna(exam_name):
                return "UNKNOWN", "UNKNOWN"

            exam_name = str(exam_name).upper().strip()

            # Match pattern like BCOM - Commerce [C032] - 1YEAR or BED - PLAIN[PLAIN] - 2SEM
            match = re.match(r'^([A-Z]+)\s*-\s*.+\[\w+\]\s*-\s*(\d+(ST|ND|RD|TH)?(YEAR|SEM))$', exam_name)
            if match:
                class_group = match.group(1).strip()
                year_or_sem = match.group(2).strip()
                return class_group, year_or_sem

            # Fallback: try to extract roman numeral patterns like II YEAR
            roman = re.search(r'\b([IVXLCDM]+)\s*(YEAR|SEM)\b', exam_name)
            if roman:
                return "UNKNOWN", roman.group(0).strip()

            return "UNKNOWN", "UNKNOWN"



        df[["Class Group", "Year"]] = df["Exam Name"].apply(lambda x: pd.Series(extract_class_group_and_year(x)))
        

        # Group definitions
        class_groups = sorted(df["Class Group"].dropna().unique())
        college_list = sorted(df["College Name"].dropna().unique())

        # Count function
        def get_counts(df, college, group, year):
            subset = df[(df["College Name"] == college) & (df["Class Group"] == group) & (df["Year"] == year)]
            total = len(subset)
            regular = len(subset[subset["Regular/Backlog"] == "REGULAR"])
            private = len(subset[subset["Regular/Backlog"] == "PRIVATE"])
            exr = len(subset[subset["Regular/Backlog"] == "EXR"])
            supp = len(subset[subset["Regular/Backlog"] == "SUPP"])
            atkt = len(subset[subset["Regular/Backlog"] == "ATKT"])
            return [total, regular, private, exr, atkt, supp]

        # Prepare output structure
        output_rows = []

        for group in class_groups:
            years = sorted(df[df["Class Group"] == group]["Year"].dropna().unique())

            # Header rows
            header_row1 = ["Class"] + [f"{group} - {year}" for year in years for _ in range(5)]
            header_row2 = ["College", "Grand Total"] + ["Total", "Regular", "Private", "EXR", "ATKT", "SUPP"] * len(years)

            block_data = []
            for college in college_list:
                row = [college]
                grand_total = 0
                for year in years:
                    t, r, p, x, a, s = get_counts(df, college, group, year)
                    row += [t, r, p, x, a, s]
                    grand_total += t
                row.insert(1, grand_total)
                block_data.append(row)

            output_rows.append(header_row1)
            output_rows.append(header_row2)
            output_rows += block_data
            output_rows.append([])

        # Final Summary Block
        output_rows.append(["College", "Total of all"])
        for college in college_list:
            total = len(df[df["College Name"] == college])
            output_rows.append([college, total])

        # Save final output
        pd.DataFrame(output_rows).to_csv(output_csv_path, index=False, header=False)
        return True, f"✅ College statistics saved to {output_csv_path}"

    except Exception as e:
        return False, f"❌ Error generating college statistics: {e}"


# New helper function to generate sequential seat numbers based on a range string (from assign_seats_app.py)
def generate_sequential_seats(seat_range_str, num_students):
    generated_seats = []
    seat_range_str = seat_range_str.strip().upper() # Normalize input

    if '-' in seat_range_str:
        start_seat_str, end_seat_str = seat_range_str.split('-')
        
        # Handle alphanumeric like "1A-60A"
        if re.match(r'^\d+[A-Z]$', start_seat_str) and re.match(r'^\d+[A-Z]$', end_seat_str):
            start_num = int(re.match(r'^(\d+)', start_seat_str).group(1))
            start_char = re.search(r'([A-Z])$', start_seat_str).group(1)
            end_num = int(re.match(r'^(\d+)', end_seat_str).group(1))
            end_char = re.search(r'([A-Z])$', end_seat_str).group(1)

            if start_char != end_char:
                raise ValueError("For alphanumeric seat ranges (e.g., 1A-60A), the alphabet part must be the same.")
            
            for i in range(start_num, end_num + 1):
                generated_seats.append(f"{i}{start_char}")
        # Handle numeric like "1-60"
        elif start_seat_str.isdigit() and end_seat_str.isdigit():
            start_num = int(start_seat_str)
            end_num = int(end_seat_str)
            for i in range(start_num, end_num + 1):
                generated_seats.append(str(i))
        else:
            raise ValueError("Invalid seat number range format. Use '1-60' or '1A-60A'.")
    elif seat_range_str.isdigit() or re.match(r'^\d+[A-Z]$', seat_range_str):
        generated_seats.append(seat_range_str)
    else:
        raise ValueError("Invalid seat number format. Use a single number, '1-60', or '1A-60A'.")

    # Return only as many seats as there are students, or all generated seats if fewer students
    return generated_seats[:num_students]


# NEW FUNCTION: Get unassigned students for a given date and shift
def get_unassigned_students_for_session(date_str, shift, sitting_plan_df, timetable_df):
    unassigned_roll_numbers_details = {} # {roll_num: {class, paper, paper_code, paper_name}}

    # 1. Filter timetable for the given date and shift
    relevant_tt_exams = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) &
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if relevant_tt_exams.empty:
        return []

    # Create a unique identifier for exams in timetable for easier matching
    relevant_tt_exams['exam_key'] = relevant_tt_exams['Class'].astype(str).str.strip().str.lower() + "_" + \
                                     relevant_tt_exams['Paper'].astype(str).str.strip() + "_" + \
                                     relevant_tt_exams['Paper Code'].astype(str).str.strip() + "_" + \
                                     relevant_tt_exams['Paper Name'].astype(str).str.strip()

    # Iterate through sitting plan to find students for these exams
    for _, sp_row in sitting_plan_df.iterrows():
        # Create exam_key for this sitting plan row
        sp_exam_key = str(sp_row['Class']).strip().lower() + "_" + \
                      str(sp_row['Paper']).strip() + "_" + \
                      str(sp_row['Paper Code']).strip() + "_" + \
                      str(sp_row['Paper Name']).strip()

        # Check if this sitting plan entry corresponds to a relevant exam session
        if sp_exam_key in relevant_tt_exams['exam_key'].values:
            room_assigned = str(sp_row['Room Number']).strip()
            
            # Check all roll numbers in this sitting plan row
            for i in range(1, 11):
                roll_col = f"Roll Number {i}"
                if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                    roll_num = str(sp_row[roll_col]).strip()
                    # If room is blank, this student is unassigned for this paper
                    if not room_assigned: # If room_assigned is an empty string
                        # Store details for display
                        unassigned_roll_numbers_details[roll_num] = {
                            'Class': str(sp_row['Class']).strip(),
                            'Paper': str(sp_row['Paper']).strip(),
                            'Paper Code': str(sp_row['Paper Code']).strip(),
                            'Paper Name': str(sp_row['Paper Name']).strip()
                        }
    
    # Convert to a list of dictionaries for display, sorted by roll number
    sorted_unassigned_list = []
    for roll, details in sorted(unassigned_roll_numbers_details.items()):
        sorted_unassigned_list.append({
            "Roll Number": roll,
            "Class": details['Class'],
            "Paper": details['Paper'],
            "Paper Code": details['Paper Code'],
            "Paper Name": details['Paper Name']
        })
    
    return sorted_unassigned_list

# NEW FUNCTION: Get summary of students by paper for a given session (assigned + unassigned)
def get_session_paper_summary(date_str, shift, sitting_plan_df, assigned_seats_df, timetable_df):
    summary_data = []

    # Filter timetable for the given date and shift
    relevant_tt_exams = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) &
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ].copy()

    if relevant_tt_exams.empty:
        return pd.DataFrame(columns=['Paper Name', 'Paper Code', 'Total Expected', 'Assigned', 'Unassigned'])

    # Iterate through each unique paper in the relevant timetable exams
    for _, tt_row in relevant_tt_exams.drop_duplicates(subset=['Paper Code', 'Paper Name']).iterrows():
        paper_code = str(tt_row['Paper Code']).strip()
        paper_name = str(tt_row['Paper Name']).strip()
        
        # Get all expected roll numbers for this specific paper (from sitting plan)
        expected_rolls = set()
        paper_sitting_rows = sitting_plan_df[sitting_plan_df['Paper Code'].astype(str).str.strip() == paper_code]
        for _, sp_row in paper_sitting_rows.iterrows():
            for i in range(1, 11):
                roll_col = f"Roll Number {i}"
                if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                    expected_rolls.add(str(sp_row[roll_col]).strip())
        
        total_expected_students = len(expected_rolls)

        # Get assigned roll numbers for this specific paper, date, and shift
        assigned_rolls_for_paper = set(
            assigned_seats_df[
                (assigned_seats_df["Paper Code"].astype(str).str.strip() == paper_code) & # Use formatted paper code
                (assigned_seats_df["Date"] == date_str) &
                (assigned_seats_df["Shift"] == shift)
            ]["Roll Number"].astype(str).tolist()
        )
        num_assigned_students = len(assigned_rolls_for_paper)

        # Calculate unassigned students
        num_unassigned_students = total_expected_students - num_assigned_students

        summary_data.append({
            'Paper Name': paper_name,
            'Paper Code': paper_code,
            'Total Expected': total_expected_students,
            'Assigned': num_assigned_students,
            'Unassigned': num_unassigned_students
        })
    
    return pd.DataFrame(summary_data)

# NEW FUNCTION: Display Room Occupancy Report
def display_room_occupancy_report(sitting_plan_df, assigned_seats_df, timetable_df):
    st.subheader("📊 Room Occupancy Report")
    st.info("View detailed occupancy for each room for a selected date and shift.")

    if sitting_plan_df.empty or timetable_df.empty:
        st.warning("Please upload 'sitting_plan.csv' and 'timetable.csv' via the Admin Panel to generate this report.")
        return

    # Date and Shift filters for the report
    report_date_options = sorted(timetable_df["Date"].dropna().unique())
    report_shift_options = sorted(timetable_df["Shift"].dropna().unique())

    if not report_date_options or not report_shift_options:
        st.info("No exam dates or shifts found in the timetable to generate a report.")
        return

    selected_report_date = st.selectbox("Select Date", report_date_options, key="room_report_date")
    selected_report_shift = st.selectbox("Select Shift", report_shift_options, key="room_report_shift")

    if st.button("Generate Room Occupancy Report"):
        # Filter sitting plan for relevant exams on this date/shift
        relevant_tt_exams = timetable_df[
            (timetable_df["Date"].astype(str).str.strip() == selected_report_date) &
            (timetable_df["Shift"].astype(str).str.strip().str.lower() == selected_report_shift.lower())
        ]

        if relevant_tt_exams.empty:
            st.info("No exams scheduled for the selected date and shift to generate room occupancy.")
            return

        # Get unique combinations of (Class, Paper, Paper Code, Paper Name) for the relevant exams
        unique_exams_in_session = relevant_tt_exams[['Class', 'Paper', 'Paper Code', 'Paper Name']].drop_duplicates()

        room_occupancy_data = []

        # Iterate through each unique room in the sitting plan that has a room number assigned
        all_rooms_in_sitting_plan = sitting_plan_df['Room Number'].dropna().astype(str).str.strip().unique()
        
        for room_num in sorted(all_rooms_in_sitting_plan):
            # Find all sitting plan entries for this room
            room_sitting_plan_entries = sitting_plan_df[sitting_plan_df['Room Number'].astype(str).str.strip() == room_num]
            
            expected_students_in_room = 0
            assigned_students_in_room = 0
            assigned_roll_numbers_list = []

            # Calculate expected students for this room for the selected session
            for _, sp_row in room_sitting_plan_entries.iterrows():
                # Create a temporary key to link sitting plan entries to unique exams in session
                sp_exam_key = str(sp_row['Class']).strip() + "_" + \
                              str(sp_row['Paper']).strip() + "_" + \
                              _format_paper_code(sp_row['Paper Code']) + "_" + \
                              str(sp_row['Paper Name']).strip()
                
                # Check if this sitting plan entry's exam is part of the current session
                is_relevant_exam = False
                for _, ue_row in unique_exams_in_session.iterrows():
                    ue_exam_key = str(ue_row['Class']).strip() + "_" + \
                                  str(ue_row['Paper']).strip() + "_" + \
                                  _format_paper_code(ue_row['Paper Code']) + "_" + \
                                  str(ue_row['Paper Name']).strip()
                    if sp_exam_key == ue_exam_key:
                        is_relevant_exam = True
                        break
                
                if is_relevant_exam:
                    for i in range(1, 11):
                        roll_col = f"Roll Number {i}"
                        if roll_col in sp_row and pd.notna(sp_row[roll_col]) and str(sp_row[roll_col]).strip() != '':
                            expected_students_in_room += 1
            
            # Get assigned students for this room for the selected session
            room_assigned_students_df = assigned_seats_df[
                (assigned_seats_df["Room Number"] == room_num) &
                (assigned_seats_df["Date"] == selected_report_date) &
                (assigned_seats_df["Shift"] == selected_report_shift)
            ]
            assigned_students_in_room = len(room_assigned_students_df)
            
            # Collect assigned roll numbers and their seats for display
            if not room_assigned_students_df.empty:
                # Sort by seat number for better readability
                def sort_seat_number_for_display(seat):
                    if isinstance(seat, str):
                        if seat.endswith('A'):
                            return (0, int(seat[:-1]))
                        elif seat.endswith('B'):
                            return (1, int(seat[:-1]))
                        elif seat.isdigit():
                            return (2, int(seat))
                    return (3, seat)
                
                room_assigned_students_df['sort_key'] = room_assigned_students_df['Seat Number'].apply(sort_seat_number_for_display)
                sorted_room_assigned = room_assigned_students_df.sort_values(by='sort_key').drop(columns=['sort_key'])

                for _, assigned_row in sorted_room_assigned.iterrows():
                    assigned_roll_numbers_list.append(
                        f"{assigned_row['Roll Number']} (Seat: {assigned_row['Seat Number']}, Paper: {assigned_row['Paper Code']})"
                    )
            
            remaining_capacity = expected_students_in_room - assigned_students_in_room
            occupancy_percentage = (assigned_students_in_room / expected_students_in_room * 100) if expected_students_in_room > 0 else 0

            room_occupancy_data.append({
                'Room Number': room_num,
                'Total Expected Students': expected_students_in_room,
                'Assigned Students': assigned_students_in_room,
                'Remaining Capacity': remaining_capacity,
                'Occupancy (%)': f"{occupancy_percentage:.2f}%",
                'Assigned Roll Numbers (Details)': ", ".join(assigned_roll_numbers_list) if assigned_roll_numbers_list else "N/A"
            })
        
        if room_occupancy_data:
            df_occupancy = pd.DataFrame(room_occupancy_data)
            st.dataframe(df_occupancy)
            
            # Optional: Download button for this report
            csv_occupancy = df_occupancy.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download Room Occupancy Report as CSV",
                data=csv_occupancy,
                file_name=f"room_occupancy_report_{selected_report_date}_{selected_report_shift}.csv",
                mime="text/csv",
            )
        else:
            st.info("No room occupancy data found for the selected date and shift.")

# NEW FUNCTION: Generate Room Chart in specified format
def generate_room_chart_report(date_str, shift, sitting_plan_df, assigned_seats_df, timetable_df):
    output_string_parts = []

    # --- Robust Checks for essential columns ---
    required_timetable_cols = ["Date", "Shift", "Time", "Class", "Paper Code", "Paper Name"]
    for col in required_timetable_cols:
        if col not in timetable_df.columns:
            return f"Error: Missing essential column '{col}' in timetable.csv. Please ensure the file is correctly formatted."

    required_assigned_seats_cols = ["Roll Number", "Paper Code", "Paper Name", "Room Number", "Seat Number", "Date", "Shift"]
    for col in required_assigned_seats_cols:
        if col not in assigned_seats_df.columns:
            return f"Error: Missing essential column '{col}' in assigned_seats.csv. Please ensure seats are assigned and the file is correctly formatted."

    # 1. Get header information from timetable
    relevant_tt_exams = timetable_df[
        (timetable_df["Date"].astype(str).str.strip() == date_str) &
        (timetable_df["Shift"].astype(str).str.strip().str.lower() == shift.lower())
    ]

    if relevant_tt_exams.empty:
        return "No exams found for the selected date and shift to generate room chart."

    # Extract common info for header (assuming they are consistent for a given date/shift)
    exam_time = relevant_tt_exams.iloc[0]["Time"].strip() if "Time" in relevant_tt_exams.columns else ""
    
    # Determine the class summary for the header
    unique_classes = relevant_tt_exams['Class'].dropna().astype(str).str.strip().unique()
    class_summary_header = ""
    if len(unique_classes) == 1:
        class_summary_header = f"{unique_classes[0]} Examination {datetime.datetime.now().year}"
    elif len(unique_classes) > 1:
        class_summary_header = f"Various Classes Examination {datetime.datetime.now().year}"
    else:
        class_summary_header = f"Examination {datetime.datetime.now().year}"

    # Static header lines
    output_string_parts.append(",,,,,,,,,\nजीवाजी विश्वविद्यालय ग्वालियर ,,,,,,,,,\n\"परीक्षा केंद्र :- शासकीय विधि महाविद्यालय, मुरेना (म. प्र.) कोड :- G107 \",,,,,,,,,\n")
    output_string_parts.append(f"{class_summary_header},,,,,,,,,\n")
    output_string_parts.append(f"Date :- ,,{date_str},,Shift :-,{shift},,Time :- ,,\n")

    # 2. Get all assigned students for the given date and shift
    assigned_students_for_session = assigned_seats_df[
        (assigned_seats_df["Date"] == date_str) &
        (assigned_seats_df["Shift"] == shift)
    ].copy()

    if assigned_students_for_session.empty:
        output_string_parts.append("\nNo students assigned seats for this date and shift.")
        return "".join(output_string_parts)

    # Merge with timetable to get full paper names and Class
    # Ensure paper codes are comparable (e.g., int vs str)
    assigned_students_for_session['Paper Code'] = assigned_students_for_session['Paper Code'].astype(str)
    timetable_df['Paper Code'] = timetable_df['Paper Code'].astype(str)

    assigned_students_for_session = pd.merge(
        assigned_students_for_session,
        timetable_df[['Paper Code', 'Paper Name', 'Class']], # Need Class for the summary line
        on='Paper Code',
        how='left',
        suffixes=('', '_tt') # Suffixes are applied only if column names are duplicated in both DFs
    )
    # Use Paper Name from timetable if available, otherwise from assigned_seats_df
    assigned_students_for_session['Paper Name'] = assigned_students_for_session['Paper Name_tt'].fillna(assigned_students_for_session['Paper Name'])
    
    # Corrected line: Access 'Class' directly, as it would not have been suffixed if not present in assigned_seats_df
    # The 'Class' column from timetable_df is merged directly if assigned_seats_df doesn't have one,
    # otherwise it would be 'Class_tt'. We need to check which one exists.
    if 'Class_tt' in assigned_students_for_session.columns:
        assigned_students_for_session['Class'] = assigned_students_for_session['Class_tt'].fillna('')
    elif 'Class' in assigned_students_for_session.columns: # Fallback if 'Class' was already in assigned_seats_df
        assigned_students_for_session['Class'] = assigned_students_for_session['Class'].fillna('')
    else:
        assigned_students_for_session['Class'] = '' # Default if neither exists, though this should be caught by earlier checks

    # Sort by Room Number, then by Seat Number
    def sort_seat_number_key(seat):
        if isinstance(seat, str):
            match_a = re.match(r'(\d+)A', seat)
            match_b = re.match(r'(\d+)B', seat)
            if match_a:
                return (0, int(match_a.group(1))) # A-seats first
            elif match_b:
                return (1, int(match_b.group(1))) # B-seats second
            elif seat.isdigit():
                return (2, int(seat)) # Numeric seats last
        return (3, seat) # Fallback for unexpected formats

    assigned_students_for_session['sort_key'] = assigned_students_for_session['Seat Number'].apply(sort_seat_number_key)
    assigned_students_for_session = assigned_students_for_session.sort_values(by=['Room Number', 'sort_key']).drop(columns=['sort_key'])

    # Group by room for output
    students_by_room = assigned_students_for_session.groupby('Room Number')

    for room_num, room_data in students_by_room:
        output_string_parts.append(f"\n,,,कक्ष  :-,{room_num}  ,,,,\n") # Room header
        
        # Get unique papers for this room and session for the "परीक्षा का नाम" line
        unique_papers_in_room = room_data[['Class', 'Paper Code', 'Paper Name']].drop_duplicates()
        
        for _, paper_row in unique_papers_in_room.iterrows():
            paper_class = str(paper_row['Class']).strip()
            paper_code = str(paper_row['Paper Code']).strip()
            paper_name = str(paper_row['Paper Name']).strip()
            
            # Count students for this specific paper in this room
            students_for_this_paper_in_room = room_data[
                (room_data['Paper Code'].astype(str).str.strip() == paper_code) &
                (room_data['Paper Name'].astype(str).str.strip() == paper_name)
            ]
            num_students_for_paper = len(students_for_this_paper_in_room)

            output_string_parts.append(
                f"परीक्षा का नाम (Class - mode - Type),,,प्रश्न पत्र (paper- paper code - paper name),,,,उत्तर पुस्तिकाएं (number of students),,\n"
                f",,,,,,,प्राप्त ,प्रयुक्त ,शेष \n"
                f"{paper_class} - Regular - Regular,,,{paper_code} - {paper_name}        ,,,,{num_students_for_paper},,\n" # Assuming Regular for now
            )
            output_string_parts.append(",,,,,,,,,\n") # Blank line

        output_string_parts.append(",,,,,,,,,\n") # Blank line
        output_string_parts.append(f",,,Total,,,,{len(room_data)},,\n") # Total for the room
        output_string_parts.append(",,,,,,,,,\n") # Blank line
        output_string_parts.append("roll number - (room number-seat number) - 20 letters of paper name,,,,,,,,,\n")

        # Now add the roll number lines
        current_line_students = []
        for _, student_row in room_data.iterrows():
            roll_num = str(student_row['Roll Number']).strip()
            room_num_display = str(student_row['Room Number']).strip()
            seat_num_display = str(student_row['Seat Number']).strip()
            paper_name_display = str(student_row['Paper Name']).strip()
            
            # Truncate paper name to first 20 characters
            truncated_paper_name = paper_name_display[:20]

            student_entry = f"{roll_num}( कक्ष-{room_num_display}-सीट-{seat_num_display})-{truncated_paper_name}"
            current_line_students.append(student_entry)

            if len(current_line_students) == 10:
                output_string_parts.append(",".join(current_line_students) + "\n")
                current_line_students = []
        
        # Add any remaining students in the last line for the room
        if current_line_students:
            output_string_parts.append(",".join(current_line_students) + "\n")
        
        output_string_parts.append("\n") # Add an extra newline between rooms

    return "".join(output_string_parts)


# Function to display the Report Panel
def display_report_panel():
    st.subheader("📊 Exam Session Reports")

    sitting_plan, timetable, assigned_seats_df = load_data() # Load assigned_seats_df here
    all_reports_df = load_cs_reports_csv()
    room_invigilators_df = load_room_invigilator_assignments() # Load room invigilators

    if all_reports_df.empty and room_invigilators_df.empty:
        st.info("No Centre Superintendent reports or invigilator assignments available yet for statistics.")
        return
    
    if sitting_plan.empty:
        st.info("Sitting plan data is required for full report statistics.")
        # We can still show basic reports if sitting_plan is empty, but attendance % will be off.

    # Initialize expected_students_df with all necessary columns from the start
    # and populate it with expected student counts from the sitting plan
    expected_students_data = []
    if not sitting_plan.empty:
        for idx, row in sitting_plan.iterrows():
            expected_count = 0
            for i in range(1, 11):
                if pd.notna(row.get(f"Roll Number {i}")) and str(row.get(f"Roll Number {i}")).strip() != '':
                    expected_count += 1
            
            expected_students_data.append({
                'Room Number': str(row['Room Number']).strip(),
                'Class': str(row['Class']).strip(), # Keep as string, lower() later
                'Paper': str(row['Paper']).strip(),   # Keep as string, lower() later
                'Paper Code': _format_paper_code(row['Paper Code']), # Use formatted paper code
                'Paper Name': str(row['Paper Name']).strip(), # Keep as string, lower() later
                'Mode': str(row.get('Mode', '')).strip(),
                'Type': str(row.get('Type', '')).strip(),
                'expected_students_count': expected_count
            })
    expected_students_df = pd.DataFrame(expected_students_data)

    # Standardize merge keys in all_reports_df
    all_reports_df['room_num'] = all_reports_df['room_num'].astype(str).str.strip()
    all_reports_df['paper_code'] = all_reports_df['paper_code'].astype(str).str.strip().str.lower()
    all_reports_df['paper_name'] = all_reports_df['paper_name'].astype(str).str.strip().str.lower()
    all_reports_df['class'] = all_reports_df['class'].astype(str).str.strip().str.lower()

    # Standardize merge keys in expected_students_df (apply .str.lower() here)
    expected_students_df['Room Number'] = expected_students_df['Room Number'].astype(str).str.strip()
    expected_students_df['Paper Code'] = expected_students_df['Paper Code'].astype(str).str.strip().str.lower()
    expected_students_df['Paper Name'] = expected_students_df['Paper Name'].astype(str).str.strip().str.lower()
    expected_students_df['Class'] = expected_students_df['Class'].astype(str).str.strip().str.lower()


    # Merge all_reports_df with expected_students_df
    # We want to keep all report entries and add expected counts where available
    merged_reports_df = pd.merge(
        all_reports_df,
        expected_students_df,
        left_on=['room_num', 'paper_code', 'paper_name', 'class'],
        right_on=['Room Number', 'Paper Code', 'Paper Name', 'Class'],
        how='left', # Use left merge to keep all reports
        suffixes=('_report', '_sp')
    )

    # Fill NaN expected_students_count with 0 for reports where no matching sitting plan entry was found
    merged_reports_df['expected_students_count'] = merged_reports_df['expected_students_count'].fillna(0).astype(int)

    # Merge with room_invigilators_df to add invigilator info
    if not room_invigilators_df.empty:
        room_invigilators_df['date'] = room_invigilators_df['date'].astype(str).str.strip()
        room_invigilators_df['shift'] = room_invigilators_df['shift'].astype(str).str.strip().str.lower()
        room_invigilators_df['room_num'] = room_invigilators_df['room_num'].astype(str).str.strip()

        merged_reports_df = pd.merge(
            merged_reports_df,
            room_invigilators_df[['date', 'shift', 'room_num', 'invigilators']],
            on=['date', 'shift', 'room_num'],
            how='left',
            suffixes=('', '_room_inv') # Suffix to avoid column name collision if 'invigilators' existed in merged_reports_df
        )
        # Fill NaN invigilators with empty list for reports where no invigilator assignment was found
        merged_reports_df['invigilators'] = merged_reports_df['invigilators'].apply(lambda x: x if isinstance(x, list) else [])

    else:
        merged_reports_df['invigilators'] = [[]] * len(merged_reports_df) # Add empty list if no invigilator data

    st.markdown("---")
    st.subheader("Overall Statistics")

    total_reports = len(merged_reports_df)
    unique_sessions = merged_reports_df['report_key'].nunique()
    total_absent = merged_reports_df['absent_roll_numbers'].apply(len).sum()
    total_ufm = merged_reports_df['ufm_roll_numbers'].apply(len).sum()
    
    # Calculate total expected students directly from the expected_students_df
    total_expected_students = expected_students_df['expected_students_count'].sum()
    
    # Calculate total present students
    total_present_students = total_expected_students - total_absent
    # Calculate total answer sheets collected
    total_answer_sheets_collected = total_present_students - total_ufm


    overall_attendance_percentage = 0
    if total_expected_students > 0:
        overall_attendance_percentage = (total_present_students / total_expected_students) * 100

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Total Reports Submitted", total_reports)
    with col2:
        st.metric("Unique Exam Sessions Reported", unique_sessions)
    with col3:
        st.metric("Total Expected Students", total_expected_students)
    with col4:
        st.metric("Total Absent Students", total_absent)
    with col5:
        st.metric("Overall Attendance (%)", f"{overall_attendance_percentage:.2f}%")
    
    col_metrics_2_1, col_metrics_2_2, col_metrics_2_3 = st.columns(3)
    with col_metrics_2_1:
        st.metric("Total Present Students", total_present_students)
    with col_metrics_2_2:
        st.metric("Total UFM Cases", total_ufm)
    with col_metrics_2_3:
        st.metric("Total Answer Sheets Collected", total_answer_sheets_collected)


    # --- Paper-wise Statistics ---
    st.markdown("---")
    st.subheader("Paper-wise Statistics")

    # Group expected students by paper
    expected_by_paper = expected_students_df.groupby(['Paper Name', 'Paper Code']).agg(
        expected_students=('expected_students_count', 'sum')
    ).reset_index()
    expected_by_paper.rename(columns={'Paper Name': 'paper_name', 'Paper Code': 'paper_code'}, inplace=True)
    expected_by_paper['paper_name'] = expected_by_paper['paper_name'].astype(str).str.strip().str.lower()
    expected_by_paper['paper_code'] = expected_by_paper['paper_code'].astype(str).str.strip().str.lower()

    # Group reported data by paper
    reported_by_paper = merged_reports_df.groupby(['paper_name', 'paper_code']).agg(
        total_absent=('absent_roll_numbers', lambda x: x.apply(len).sum()),
        total_ufm=('ufm_roll_numbers', lambda x: x.apply(len).sum())
    ).reset_index()

    # Merge expected and reported data
    paper_stats = pd.merge(
        expected_by_paper,
        reported_by_paper,
        on=['paper_name', 'paper_code'],
        how='left' # Keep all papers from expected_students_df
    )

    # Fill NaN values for absent/ufm with 0 where no reports exist
    paper_stats['total_absent'] = paper_stats['total_absent'].fillna(0).astype(int)
    paper_stats['total_ufm'] = paper_stats['total_ufm'].fillna(0).astype(int)

    paper_stats['total_present'] = paper_stats['expected_students'] - paper_stats['total_absent']
    paper_stats['total_answer_sheets_collected'] = paper_stats['total_present'] - paper_stats['total_ufm']
    paper_stats['attendance_percentage'] = paper_stats.apply(
        lambda row: (row['total_present'] / row['expected_students'] * 100) if row['expected_students'] > 0 else 0,
        axis=1
    )
    paper_stats['attendance_percentage'] = paper_stats['attendance_percentage'].map('{:.2f}%'.format)

    # Rename columns for display
    paper_stats.rename(columns={
        'paper_name': 'Paper Name',
        'paper_code': 'Paper Code',
        'expected_students': 'Expected Students',
        'total_absent': 'Absent Students',
        'total_present': 'Present Students',
        'total_ufm': 'UFM Cases',
        'total_answer_sheets_collected': 'Answer Sheets Collected',
        'attendance_percentage': 'Attendance (%)'
    }, inplace=True)

    st.dataframe(paper_stats[['Paper Name', 'Paper Code', 'Expected Students', 'Present Students', 'Absent Students', 'UFM Cases', 'Answer Sheets Collected', 'Attendance (%)']])


    # --- Student Type-wise Statistics ---
    st.markdown("---")
    st.subheader("Student Type-wise Statistics")

    # Group expected students by Class, Mode, Type
    expected_by_type = expected_students_df.groupby(['Class', 'Mode', 'Type']).agg(
        expected_students=('expected_students_count', 'sum')
    ).reset_index()
    expected_by_type.rename(columns={'Class': 'Class_sp', 'Mode': 'Mode_sp', 'Type': 'Type_sp'}, inplace=True)


    # Group reported data by Class, Mode, Type (from the merged_reports_df which has _sp suffixes)
    # Ensure these columns exist before grouping, as they come from the sitting plan side of the merge
    required_type_cols_for_grouping = ['Class_sp', 'Mode_sp', 'Type_sp']
    
    # Filter merged_reports_df to ensure we only consider rows where type info is available
    if all(col in merged_reports_df.columns for col in required_type_cols_for_grouping):
        reported_by_type_df = merged_reports_df.dropna(subset=required_type_cols_for_grouping).copy()

        if not reported_by_type_df.empty:
            reported_by_type = reported_by_type_df.groupby(required_type_cols_for_grouping).agg(
                total_absent=('absent_roll_numbers', lambda x: x.apply(len).sum()),
                total_ufm=('ufm_roll_numbers', lambda x: x.apply(len).sum())
            ).reset_index()

            # Merge expected and reported data
            type_stats = pd.merge(
                expected_by_type,
                reported_by_type,
                on=required_type_cols_for_grouping,
                how='left' # Keep all types from expected_students_df
            )

            # Fill NaN values for absent/ufm with 0 where no reports exist
            type_stats['total_absent'] = type_stats['total_absent'].fillna(0).astype(int)
            type_stats['total_ufm'] = type_stats['total_ufm'].fillna(0).astype(int)

            type_stats['total_present'] = type_stats['expected_students'] - type_stats['total_absent']
            type_stats['total_answer_sheets_collected'] = type_stats['total_present'] - type_stats['total_ufm']
            type_stats['attendance_percentage'] = type_stats.apply(
                lambda row: (row['total_present'] / row['expected_students'] * 100) if row['expected_students'] > 0 else 0,
                axis=1
            )
            type_stats['attendance_percentage'] = type_stats['attendance_percentage'].map('{:.2f}%'.format)

            # Rename columns for display
            type_stats.rename(columns={
                'Class_sp': 'Class',
                'Mode_sp': 'Mode',
                'Type_sp': 'Type',
                'expected_students': 'Expected Students',
                'total_absent': 'Absent Students',
                'total_present': 'Present Students',
                'total_ufm': 'UFM Cases',
                'total_answer_sheets_collected': 'Answer Sheets Collected',
                'attendance_percentage': 'Attendance (%)'
            }, inplace=True)

            st.dataframe(type_stats[['Class', 'Mode', 'Type', 'Expected Students', 'Present Students', 'Absent Students', 'UFM Cases', 'Answer Sheets Collected', 'Attendance (%)']])
        else:
            st.info("No student type data available in reports for statistics after filtering.")
    else:
        st.info("Required student type columns (Class, Mode, Type) are not available in the merged reports for statistics.")


    st.markdown("---")
    st.subheader("Filter and View Reports")

    # Filters
    unique_dates = sorted(merged_reports_df['date'].unique())
    unique_shifts = sorted(merged_reports_df['shift'].unique())
    unique_rooms = sorted(merged_reports_df['room_num'].unique())
    unique_papers = sorted(merged_reports_df['paper_name'].unique())

    filter_date = st.selectbox("Filter by Date", ["All"] + unique_dates, key="report_filter_date")
    filter_shift = st.selectbox("Filter by Shift", ["All"] + unique_shifts, key="report_filter_shift")
    filter_room = st.selectbox("Filter by Room Number", ["All"] + unique_rooms, key="report_filter_room")
    filter_paper = st.selectbox("Filter by Paper Name", ["All"] + unique_papers, key="report_filter_paper")

    filtered_reports_df = merged_reports_df.copy()

    if filter_date != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['date'] == filter_date]
    if filter_shift != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['shift'] == filter_shift]
    if filter_room != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['room_num'] == filter_room]
    if filter_paper != "All":
        filtered_reports_df = filtered_reports_df[filtered_reports_df['paper_name'] == filter_paper]

    if filtered_reports_df.empty:
        st.info("No reports match the selected filters.")
    else:
        st.markdown("---")
        st.subheader("Filtered Reports Summary")
        st.dataframe(filtered_reports_df[[
            'date', 'shift', 'room_num', 'paper_code', 'paper_name', 'invigilators', # 'invigilators' is now from merge
            'absent_roll_numbers', 'ufm_roll_numbers'
        ]].rename(columns={
            'date': 'Date', 'shift': 'Shift', 'room_num': 'Room',
            'paper_code': 'Paper Code', 'paper_name': 'Paper Name',
            'invigilators': 'Invigilators',
            'absent_roll_numbers': 'Absent Roll Numbers',
            'ufm_roll_numbers': 'UFM Roll Numbers'
        }))

        st.markdown("---")
        st.subheader("Detailed Absentee List (Filtered)")
        absent_list_data = []
        for _, row in filtered_reports_df.iterrows():
            for roll in row['absent_roll_numbers']:
                absent_list_data.append({
                    'Date': row['date'],
                    'Shift': row['shift'],
                    'Room': row['room_num'],
                    'Paper Code': row['paper_code'],
                    'Paper Name': row['paper_name'],
                    'Absent Roll Number': roll
                })
        
        if absent_list_data:
            df_absent = pd.DataFrame(absent_list_data)
            st.dataframe(df_absent)
            
            # Download Absentee List
            csv_absent = df_absent.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download Absentee List as CSV",
                data=csv_absent,
                file_name=f"absent_list_{filter_date}_{filter_shift}.csv",
                mime="text/csv",
            )
        else:
            st.info("No absent students in the filtered reports.")

        st.markdown("---")
        st.subheader("Detailed UFM List (Filtered)")
        ufm_list_data = []
        for _, row in filtered_reports_df.iterrows():
            for roll in row['ufm_roll_numbers']:
                ufm_list_data.append({
                    'Date': row['date'],
                    'Shift': row['shift'],
                    'Room': row['room_num'],
                    'Paper Code': row['paper_code'],
                    'Paper Name': row['paper_name'],
                    'UFM Roll Number': roll
                })
        
        if ufm_list_data:
            df_ufm = pd.DataFrame(ufm_list_data)
            st.dataframe(df_ufm)

            # Download UFM List
            csv_ufm = df_ufm.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download UFM List as CSV",
                data=csv_ufm,
                file_name=f"ufm_list_{filter_date}_{filter_shift}.csv",
                mime="text/csv",
            )
        else:
            st.info("No UFM cases in the filtered reports.")
# --- Updated Remuneration Calculation Functions (from bill.py) ---
def calculate_remuneration(shift_assignments_df, room_invigilator_assignments_df, timetable_df, assigned_seats_df,
                            manual_rates, prep_closing_assignments, holiday_dates, selected_classes_for_bill):
    """
    Calculates the remuneration for all team members based on assignments and rules,
    including individually selected preparation and closing days and holiday conveyance allowance.
    
    Updated Rules:
    1. Person gets conveyance in evening shift of selected exam (if also worked in evening shift of selected/non-selected exam)
    2. Person doesn't get conveyance in morning shift of selected exam (if also worked in evening shift of selected/non-selected exam)
    3. Senior CS gets daily remuneration if worked in either shift of selected exam in bill of selected exam
    4. Senior CS doesn't get daily remuneration if worked in morning shift of selected exam and evening shift of non-selected exam in bill of selected exam
    5. Senior CS gets daily remuneration if worked in both shifts of selected exam in bill of selected exam
    """
    remuneration_data_detailed_raw = []
    
    # Define remuneration rules and their base rates
    remuneration_rules = {
        'senior_center_superintendent': {'role_display': 'Senior Center Superintendent', 'rate': manual_rates['senior_center_superintendent_rate'], 'unit': 'day', 'eligible_prep_close': True, 'exam_conveyance': False},
        'center_superintendent': {'role_display': 'Center Superintendent', 'rate': manual_rates['center_superintendent_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'assistant_center_superintendent': {'role_display': 'Assistant Center Superintendent', 'rate': manual_rates['assistant_center_superintendent_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'permanent_invigilator': {'role_display': 'Permanent Invigilator', 'rate': manual_rates['permanent_invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': True, 'exam_conveyance': True},
        'assistant_permanent_invigilator': {'role_display': 'Assistant Permanent Invigilator', 'rate': manual_rates['assistant_permanent_invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': False, 'exam_conveyance': True},
        'invigilator': {'role_display': 'Invigilator', 'rate': manual_rates['invigilator_rate'], 'unit': 'shift', 'eligible_prep_close': False, 'exam_conveyance': True},
    }

    # Define rates for Class 3 and 4 workers (per student, applied to total students)
    class_worker_rates = {
        'class_3_worker': {'role_display': 'Class 3 Worker', 'rate_per_student': manual_rates['class_3_worker_rate_per_student']},
        'class_4_worker': {'role_display': 'Class 4 Worker', 'rate_per_student': manual_rates['class_4_worker_rate_per_student']},
    }

    # Create a unified list of all assigned personnel and their roles for easier iteration
    unified_assignments = []

    # Collect unique Class 3 and Class 4 workers for overall remuneration calculation
    unique_class_3_workers = set()
    unique_class_4_workers = set()

    # Process shift assignments (all shifts, for conveyance and Senior CS daily rate)
    for index, row in shift_assignments_df.iterrows():
        current_date = row['date']
        current_shift = row['shift']

        for role_col in remuneration_rules.keys():
            if role_col in row and isinstance(row[role_col], list):
                for person in row[role_col]:
                    unified_assignments.append({
                        'Name': person,
                        'Role_Key': role_col,
                        'Date': current_date,
                        'Shift': current_shift,
                        'Source': 'shift_assignments'
                    })
        
        # Collect unique Class 3 and 4 workers
        if 'class_3_worker' in row and isinstance(row['class_3_worker'], list):
            unique_class_3_workers.update(row['class_3_worker'])
        if 'class_4_worker' in row and isinstance(row['class_4_worker'], list):
            unique_class_4_workers.update(row['class_4_worker'])

    # Process room invigilator assignments (all shifts, for conveyance)
    for index, row in room_invigilator_assignments_df.iterrows():
        current_date = row['date']
        current_shift = row['shift']
        invigilators_list = row['invigilators']

        for invigilator in invigilators_list:
            is_assigned_higher_role = False
            for assignment in unified_assignments:
                if (assignment['Name'] == invigilator and
                    assignment['Date'] == current_date and
                    assignment['Shift'] == current_shift and
                    assignment['Role_Key'] != 'invigilator'):
                    is_assigned_higher_role = True
                    break
            
            if not is_assigned_higher_role:
                unified_assignments.append({
                    'Name': invigilator,
                    'Role_Key': 'invigilator',
                    'Date': current_date,
                    'Shift': current_shift,
                    'Source': 'room_invigilator_assignments'
                })

    # Convert to DataFrame for easier processing
    df_assignments = pd.DataFrame(unified_assignments)
    
    # Create mapping from (Date, Shift) to Classes for exam classification
    session_classes_map = {}
    for _, tt_row in timetable_df.iterrows():
        date_shift_key = (str(tt_row['Date']), str(tt_row['Shift']))
        if date_shift_key not in session_classes_map:
            session_classes_map[date_shift_key] = set()
        session_classes_map[date_shift_key].add(str(tt_row['Class']).strip())

    # NEW CONVEYANCE LOGIC: Check if person worked evening shift of any exam (selected or non-selected)
    evening_shift_workers = {}
    if not df_assignments.empty:
        df_assignments['Date_dt'] = pd.to_datetime(df_assignments['Date'], format='%d-%m-%Y', errors='coerce')
        evening_workers = df_assignments[df_assignments['Shift'] == 'Evening']
        for _, row in evening_workers.iterrows():
            name = row['Name']
            if name not in evening_shift_workers:
                evening_shift_workers[name] = set()
            evening_shift_workers[name].add((row['Date'], row['Role_Key']))

    # Now calculate remuneration for all entries in unified_assignments
    for assignment in unified_assignments:
        name = assignment['Name']
        role_key = assignment['Role_Key']
        date = assignment['Date']
        shift = assignment['Shift']

        # Get classes for this session
        session_classes = list(session_classes_map.get((date, shift), set()))
        is_selected_exam = any(cls in selected_classes_for_bill for cls in [c.strip() for c in session_classes]) if selected_classes_for_bill else True

        # Base remuneration for the shift
        base_rem_for_shift = remuneration_rules[role_key]['rate']
        
        # NEW CONVEYANCE LOGIC IMPLEMENTATION
        conveyance = 0
        if remuneration_rules[role_key]['exam_conveyance']:
            # Rule 1: Person gets conveyance in evening shift of selected exam (if also worked in evening shift)
            if shift == 'Evening' and is_selected_exam and name in evening_shift_workers:
                conveyance = manual_rates['conveyance_rate']
            
            # Rule 2: Person doesn't get conveyance in morning shift of selected exam (if also worked in evening shift)
            elif shift == 'Morning' and is_selected_exam and name in evening_shift_workers:
                conveyance = 0
            
            # Default conveyance for other cases (non-selected exams, etc.)
            elif shift == 'Evening' and not name in evening_shift_workers:
                conveyance = manual_rates['conveyance_rate']

        remuneration_data_detailed_raw.append({
            'Name': name,
            'Role_Key': role_key,
            'Role_Display': remuneration_rules[role_key]['role_display'],
            'Date': date,
            'Shift': shift,
            'Base_Remuneration_Per_Shift_Unfiltered': base_rem_for_shift,
            'Conveyance': conveyance,
            'Is_Selected_Exam': is_selected_exam,
            'Classes_in_Session': session_classes,
        })
    
    df_detailed_remuneration = pd.DataFrame(remuneration_data_detailed_raw)

    # --- Generate Individual Bills ---
    individual_bills = []
    unique_person_roles = df_detailed_remuneration[['Name', 'Role_Display', 'Role_Key']].drop_duplicates()

    for idx, row in unique_person_roles.iterrows():
        name = row['Name']
        role_display = row['Role_Display']
        role_key = row['Role_Key']

        person_data = df_detailed_remuneration[
            (df_detailed_remuneration['Name'] == name) &
            (df_detailed_remuneration['Role_Display'] == role_display)
        ].copy()
        
        # UPDATED SENIOR CS REMUNERATION LOGIC
        if role_key == 'senior_center_superintendent':
            # Get person's work pattern
            selected_shifts = person_data[person_data['Is_Selected_Exam'] == True]
            non_selected_shifts = person_data[person_data['Is_Selected_Exam'] == False]
            
            # Group by date to check daily patterns
            selected_dates = {}
            non_selected_dates = {}
            
            for _, shift_row in selected_shifts.iterrows():
                date = shift_row['Date']
                if date not in selected_dates:
                    selected_dates[date] = []
                selected_dates[date].append(shift_row['Shift'])
            
            for _, shift_row in non_selected_shifts.iterrows():
                date = shift_row['Date']
                if date not in non_selected_dates:
                    non_selected_dates[date] = []
                non_selected_dates[date].append(shift_row['Shift'])
            
            # Apply Senior CS rules
            eligible_days = set()
            
            for date, shifts in selected_dates.items():
                # Rule 3: Gets daily remuneration if worked in either shift of selected exam
                # Rule 5: Gets daily remuneration if worked in both shifts of selected exam
                if 'Morning' in shifts or 'Evening' in shifts:
                    eligible_days.add(date)
            
            # Rule 4: Remove days where worked morning shift of selected exam and evening shift of non-selected exam
            for date in list(eligible_days):
                if (date in selected_dates and 'Morning' in selected_dates[date] and
                    date in non_selected_dates and 'Evening' in non_selected_dates[date] and
                    date not in selected_dates or 'Evening' not in selected_dates.get(date, [])):
                    eligible_days.remove(date)
            
            filtered_person_data = person_data[person_data['Date'].isin(eligible_days)]
        else:
            # Filter other roles by selected classes for duties on exam days
            if selected_classes_for_bill:
                filtered_person_data = person_data[person_data['Is_Selected_Exam'] == True].copy()
            else:
                filtered_person_data = person_data.copy()

        # Group dates by month for display for morning shifts (filtered data)
        duty_dates_morning_str = ""
        morning_shifts_df = filtered_person_data[filtered_person_data['Shift'] == 'Morning']
        if not morning_shifts_df.empty:
            morning_shifts_df['Date_dt'] = pd.to_datetime(morning_shifts_df['Date'], format='%d-%m-%Y', errors='coerce')
            morning_shifts_df = morning_shifts_df.sort_values(by='Date_dt')
            grouped_dates_morning = morning_shifts_df.groupby(morning_shifts_df['Date_dt'].dt.to_period('M'))['Date_dt'].apply(lambda x: sorted(x.dt.day.tolist()))
            date_parts = []
            for period, days in grouped_dates_morning.items():
                month_name = period.strftime('%b')
                days_str = ", ".join(map(str, days))
                date_parts.append(f"{month_name} - {days_str}")
            if date_parts:
                duty_dates_morning_str = ", ".join(date_parts) + f" {morning_shifts_df['Date_dt'].min().year}"

        # Group dates by month for display for evening shifts (filtered data)
        duty_dates_evening_str = ""
        evening_shifts_df = filtered_person_data[filtered_person_data['Shift'] == 'Evening']
        if not evening_shifts_df.empty:
            evening_shifts_df['Date_dt'] = pd.to_datetime(evening_shifts_df['Date'], format='%d-%m-%Y', errors='coerce')
            evening_shifts_df = evening_shifts_df.sort_values(by='Date_dt')
            grouped_dates_evening = evening_shifts_df.groupby(evening_shifts_df['Date_dt'].dt.to_period('M'))['Date_dt'].apply(lambda x: sorted(x.dt.day.tolist()))
            date_parts = []
            for period, days in grouped_dates_evening.items():
                month_name = period.strftime('%b')
                days_str = ", ".join(map(str, days))
                date_parts.append(f"{month_name} - {days_str}")
            if date_parts:
                duty_dates_evening_str = ", ".join(date_parts) + f" {evening_shifts_df['Date_dt'].min().year}"

        total_morning_shifts = len(morning_shifts_df)
        total_evening_shifts = len(evening_shifts_df)
        total_shifts = total_morning_shifts + total_evening_shifts
        rate_in_rs = remuneration_rules[role_key]['rate'] if role_key in remuneration_rules else 0

        # Calculate base remuneration
        total_base_remuneration = 0
        if role_key == 'senior_center_superintendent':
            # Senior CS: Rs. per day based on eligible days
            unique_dates = filtered_person_data['Date'].nunique()
            total_base_remuneration = unique_dates * rate_in_rs
        else:
            # Other roles (per shift): Filter base remuneration by selected classes
            total_base_remuneration = filtered_person_data['Base_Remuneration_Per_Shift_Unfiltered'].sum()
        
        # Conveyance is calculated based on all shifts (not filtered by selected classes)
        total_conveyance = person_data['Conveyance'].sum() 
        
        # Calculate preparation and closing day remuneration - ROLE SPECIFIC
        total_prep_remuneration = 0
        total_closing_remuneration = 0
        total_holiday_conveyance = 0
        
        # Check if this person is eligible and assigned prep/closing days for THIS SPECIFIC ROLE
        if remuneration_rules[role_key]['eligible_prep_close']:
            person_assignments = prep_closing_assignments.get(name, {})
            assigned_role = person_assignments.get('role')
            
            # Only apply prep/closing if the assignment matches current role
            if assigned_role == role_key:
                # Preparation days
                prep_days = person_assignments.get('prep_days', [])
                total_prep_remuneration = len(prep_days) * rate_in_rs
                
                # Closing days
                closing_days = person_assignments.get('closing_days', [])
                total_closing_remuneration = len(closing_days) * rate_in_rs
                
                # Holiday conveyance for prep/closing days that are holidays
                all_assigned_days = prep_days + closing_days
                holiday_assigned_days = [day for day in all_assigned_days if day in holiday_dates]
                total_holiday_conveyance = len(holiday_assigned_days) * manual_rates['holiday_conveyance_allowance_rate']

        grand_total_amount = total_base_remuneration + total_conveyance + total_prep_remuneration + total_closing_remuneration + total_holiday_conveyance

        individual_bills.append({
            'SN': len(individual_bills) + 1,
            'Name (with role)': f"{name} ({role_display})",
            'Duty dates of selected class exam Shift (morning)': duty_dates_morning_str,
            'Duty dates of selected class exam Shift (evening)': duty_dates_evening_str,
            'Total shifts of selected class exams (morning/evening)': total_shifts,
            'Rate in Rs': rate_in_rs,
            'Total Remuneration in Rs': total_base_remuneration,
            'Total Conveyance (in evening shift)': total_conveyance,
            'Preparation Day Remuneration': total_prep_remuneration,
            'Closing Day Remuneration': total_closing_remuneration,
            'Total Holiday Conveyance Added': total_holiday_conveyance,
            'Total amount in Rs': grand_total_amount,
            'Signature': ''
        })

    df_individual_bills = pd.DataFrame(individual_bills)

    # --- Generate Role-wise Summary Matrix ---
    role_summary_matrix = []
    
    for role_key, rule in remuneration_rules.items():
        role_df = df_detailed_remuneration[df_detailed_remuneration['Role_Key'] == role_key]
        
        if not role_df.empty:
            total_shifts_count = len(role_df)
            
            # UPDATED REMUNERATION CALCULATION FOR SUMMARY
            total_base_rem = 0
            if role_key == 'senior_center_superintendent':
                # Senior CS: Apply the same logic as individual bills
                for name in role_df['Name'].unique():
                    person_data = role_df[role_df['Name'] == name]
                    
                    # Apply Senior CS rules for this person
                    selected_shifts = person_data[person_data['Is_Selected_Exam'] == True]
                    non_selected_shifts = person_data[person_data['Is_Selected_Exam'] == False]
                    
                    selected_dates = {}
                    non_selected_dates = {}
                    
                    for _, shift_row in selected_shifts.iterrows():
                        date = shift_row['Date']
                        if date not in selected_dates:
                            selected_dates[date] = []
                        selected_dates[date].append(shift_row['Shift'])
                    
                    for _, shift_row in non_selected_shifts.iterrows():
                        date = shift_row['Date']
                        if date not in non_selected_dates:
                            non_selected_dates[date] = []
                        non_selected_dates[date].append(shift_row['Shift'])
                    
                    eligible_days = set()
                    
                    for date, shifts in selected_dates.items():
                        if 'Morning' in shifts or 'Evening' in shifts:
                            eligible_days.add(date)
                    
                    for date in list(eligible_days):
                        if (date in selected_dates and 'Morning' in selected_dates[date] and
                            date in non_selected_dates and 'Evening' in non_selected_dates[date] and
                            (date not in selected_dates or 'Evening' not in selected_dates.get(date, []))):
                            eligible_days.remove(date)
                    
                    total_base_rem += len(eligible_days) * rule['rate']
            else:
                # Other roles (per shift): Filter base remuneration by selected classes for summary
                for _, r_row in role_df.iterrows():
                    if r_row['Is_Selected_Exam']:
                        total_base_rem += r_row['Base_Remuneration_Per_Shift_Unfiltered']

            total_conveyance = role_df['Conveyance'].sum() # Conveyance is NOT filtered by selected classes
            
            # Calculate total prep/closing remuneration for this role - ROLE SPECIFIC
            total_prep_added = 0
            total_closing_added = 0
            total_holiday_conveyance = 0
            
            if rule['eligible_prep_close']:
                # Only count prep/closing for people assigned to THIS specific role
                for name, person_assignments in prep_closing_assignments.items():
                    assigned_role = person_assignments.get('role')
                    if assigned_role == role_key:
                        prep_days = person_assignments.get('prep_days', [])
                        closing_days = person_assignments.get('closing_days', [])
                        total_prep_added += len(prep_days) * rule['rate']
                        total_closing_added += len(closing_days) * rule['rate']
                        
                        # Holiday conveyance
                        all_assigned_days = prep_days + closing_days
                        holiday_assigned_days = [day for day in all_assigned_days if day in holiday_dates]
                        total_holiday_conveyance += len(holiday_assigned_days) * manual_rates['holiday_conveyance_allowance_rate']
            
            grand_total = total_base_rem + total_conveyance + total_prep_added + total_closing_added + total_holiday_conveyance

            role_summary_matrix.append({
                'SN': len(role_summary_matrix) + 1,
                'Name (with role)': rule['role_display'],
                'Duty dates': '',
                'Shift (morning/evening)': '',
                'Total shifts (morning/evening)': total_shifts_count,
                'Rate in Rs': rule['rate'],
                'Total Remuneration in Rs': total_base_rem,
                'Total Conveyance (in evening shift)': total_conveyance,
                'Preparation Day Remuneration': total_prep_added,
                'Closing Day Remuneration': total_closing_added,
                'Total Holiday Conveyance Added': total_holiday_conveyance,
                'Total amount in Rs': grand_total,
                'Signature': ''
            })
    
    df_role_summary_matrix = pd.DataFrame(role_summary_matrix)

    # --- Generate Class 3 and Class 4 Worker Bills ---
    class_3_4_final_bills = []

    # --- UPDATED LOGIC FOR TOTAL STUDENTS FOR CLASS WORKERS ---
    # Class 3/4 worker remuneration is based on total unique students, filtered by selected classes
    if selected_classes_for_bill:
        # Get paper codes for the selected classes from the timetable
        papers_for_selected_classes = timetable_df[timetable_df['Class'].isin(selected_classes_for_bill)]['Paper Code'].unique()

        # Filter assigned seats based on these paper codes
        filtered_assigned_seats = assigned_seats_df[assigned_seats_df['Paper Code'].isin(papers_for_selected_classes)]

        # Calculate total unique students from the filtered list
        total_students_for_class_workers = filtered_assigned_seats['Roll Number'].nunique()
    else:
        # If no specific class is selected, consider all students across all exams
        total_students_for_class_workers = assigned_seats_df['Roll Number'].nunique()

    # Calculate remuneration for Class 3 workers
    if unique_class_3_workers:
        class_3_total_fixed_amount = total_students_for_class_workers * class_worker_rates['class_3_worker']['rate_per_student']
        num_class_3_workers = len(unique_class_3_workers)
        rem_per_class_3_worker = class_3_total_fixed_amount / num_class_3_workers if num_class_3_workers > 0 else 0

        for sn, worker_name in enumerate(sorted(list(unique_class_3_workers))):
            class_3_4_final_bills.append({
                'S.N.': len(class_3_4_final_bills) + 1,
                'Name': worker_name,
                'Role': class_worker_rates['class_3_worker']['role_display'],
                'Total Students (Center-wide)': total_students_for_class_workers,
                'Rate per Student (for category)': class_worker_rates['class_3_worker']['rate_per_student'],
                'Total Remuneration for Category (Rs.)': class_3_total_fixed_amount,
                'Number of Workers in Category': num_class_3_workers,
                'Remuneration per Worker in Rs.': rem_per_class_3_worker,
                'Signature of Receiver': ''
            })

    # Calculate remuneration for Class 4 workers
    if unique_class_4_workers:
        class_4_total_fixed_amount = total_students_for_class_workers * class_worker_rates['class_4_worker']['rate_per_student']
        num_class_4_workers = len(unique_class_4_workers)
        rem_per_class_4_worker = class_4_total_fixed_amount / num_class_4_workers if num_class_4_workers > 0 else 0

        for sn, worker_name in enumerate(sorted(list(unique_class_4_workers))):
            class_3_4_final_bills.append({
                'S.N.': len(class_3_4_final_bills) + 1,
                'Name': worker_name,
                'Role': class_worker_rates['class_4_worker']['role_display'],
                'Total Students (Center-wide)': total_students_for_class_workers,
                'Rate per Student (for category)': class_worker_rates['class_4_worker']['rate_per_student'],
                'Total Remuneration for Category (Rs.)': class_4_total_fixed_amount,
                'Number of Workers in Category': num_class_4_workers,
                'Remuneration per Worker in Rs.': rem_per_class_4_worker,
                'Signature of Receiver': ''
            })

    df_class_3_4_final_bills = pd.DataFrame(class_3_4_final_bills)
    
    return df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills

def add_total_row(df):
    """Add a total row to the dataframe"""
    if df.empty:
        return df
    
    total_row = {}
    for col in df.columns:
        if col in ['SN', 'S.N.']:
            total_row[col] = 'TOTAL'
        elif col in ['Name (with role)', 'Name', 'Role', 'Duty dates', 'Shift (morning/evening)', 'Signature', 'Signature of Receiver',
                     'Duty dates of selected class exam Shift (morning)', 'Duty dates of selected class exam Shift (evening)']:
            total_row[col] = ''
        elif df[col].dtype in ['int64', 'float64']:
            total_row[col] = df[col].sum()
        else:
            total_row[col] = ''
    
    total_df = pd.DataFrame([total_row])
    return pd.concat([df, total_df], ignore_index=True)

def save_bills_to_excel(individual_bills_df, role_summary_df, class_workers_df, filename="remuneration_bills.xlsx"):
    """
    Saves the three remuneration dataframes into a single Excel file with multiple sheets.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not individual_bills_df.empty:
            individual_bills_df.to_excel(writer, sheet_name='Individual Bills', index=False)
            # Auto-adjust column width for individual bills
            worksheet = writer.sheets['Individual Bills']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter # Get the column name
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width

        if not role_summary_df.empty:
            role_summary_df.to_excel(writer, sheet_name='Role Summary', index=False)
            # Auto-adjust column width for role summary
            worksheet = writer.sheets['Role Summary']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width

        if not class_workers_df.empty:
            class_workers_df.to_excel(writer, sheet_name='Class Workers', index=False)
            # Auto-adjust column width for class workers
            worksheet = writer.sheets['Class Workers']
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_name].width = adjusted_width
    
    output.seek(0)
    return output, filename

# Main app
st.title("Government Law College, Morena (M.P.) Examination Management System")

menu = st.radio("Select Module", ["Student View", "Admin Panel", "Centre Superintendent Panel"])

if menu == "Student View":
    sitting_plan, timetable, assigned_seats_df = load_data() # Load assigned_seats_df here

    # Check if dataframes are empty, indicating files were not loaded
    if sitting_plan.empty or timetable.empty:
        st.warning("Sitting plan or timetable data not found. Please upload them via the Admin Panel for full functionality.")
    
    option = st.radio("Choose Search Option:", [
        "Search by Roll Number and Date",
        "Get Full Exam Schedule by Roll Number",
        "View Full Timetable"
    ])

    if option == "Search by Roll Number and Date":
        roll = st.text_input("Enter Roll Number", max_chars=9)
        date_input = st.date_input("Enter Exam Date", value=datetime.date.today())
        if st.button("Search"):
            if sitting_plan.empty or timetable.empty:
                st.warning("Sitting plan or timetable data is missing. Please upload them via the Admin Panel to search.")
            else:
                results = get_sitting_details(roll, date_input.strftime('%d-%m-%Y'), sitting_plan, timetable)
                if results:
                    st.success(f"Found {len(results)} exam(s) for Roll Number {roll} on {date_input.strftime('%d-%m-%Y')}:")
                    for i, result in enumerate(results):
                        st.markdown(f"---")
                        st.subheader(f"Exam {i+1}")
                        st.write(f"**Room Number:** {result['Room Number']}") # Display as string
                        st.write(f"**🪑 Seat Number:** {result['Seat Number']}") # Display as string
                        st.write(f"**📚 Paper:** {result['Paper']} - {result['Paper Name']} - ({result['Paper Code']})")
                        st.write(f"**🏫 Class:** {result['Class']}")
                        st.write(f"**🎓 Student type:** {result['Mode']} - {result['Type']}")
                        st.write(f"**🕐 Shift:** {result['Shift']}, **📅 Date:** {result['Date']}")
                else:
                    st.warning("No data found for the given inputs.")

    elif option == "Get Full Exam Schedule by Roll Number":
        roll = st.text_input("Enter Roll Number")
        if st.button("Get Schedule"):
            if sitting_plan.empty or timetable.empty:
                st.warning("Sitting plan or timetable data is missing. Please upload them via the Admin Panel to get schedule.")
            else:
                schedule = pd.DataFrame(get_all_exams(roll, sitting_plan, timetable))
                if not schedule.empty:
                    schedule['Date_dt'] = pd.to_datetime(schedule['Date'], format='%d-%m-%Y', errors='coerce')
                    schedule = schedule.sort_values(by="Date_dt").drop(columns=['Date_dt'])
                    st.write(schedule)
                else:
                    st.warning("No exam records found for this roll number.")
    
    elif option == "View Full Timetable":
        st.subheader("Full Examination Timetable")
        if timetable.empty:
            st.warning("Timetable data is missing. Please upload it via the Admin Panel.")
        else:
            st.dataframe(timetable)


elif menu == "Admin Panel":
    st.subheader("🔐 Admin Login")
    if admin_login():
        st.success("Login successful!")
        
        # Load data here, inside the successful login block
        sitting_plan, timetable, assigned_seats_df = load_data()

        
        
        st.markdown("---")
        st.subheader("Current Data Previews")
        col_sp, col_tt, col_assigned = st.columns(3) # Added a column for assigned_seats
        with col_sp:
            st.write(f"**{SITTING_PLAN_FILE}**")
            if not sitting_plan.empty:
                st.dataframe(sitting_plan)
            else:
                st.info("No sitting plan data loaded.")
        with col_tt:
            st.write(f"**{TIMETABLE_FILE}**")
            if not timetable.empty:
                st.dataframe(timetable)
            else:
                st.info("No timetable data loaded.")
        with col_assigned: # Display assigned_seats.csv
            st.write(f"**{ASSIGNED_SEATS_FILE}**")
            if not assigned_seats_df.empty:
                st.dataframe(assigned_seats_df)
            else:
                st.info("No assigned seats data loaded.")

        st.markdown("---") # Separator

        # Admin Panel Options
        admin_option = st.radio("Select Admin Task:", [
            "Get All Students for Date & Shift (Room Wise)",
            "Get All Students for Date & Shift (Roll Number Wise)",
            "Update Timetable Details",
            "Assign Rooms & Seats to Students", # Renamed option
            "Room Occupancy Report", # New option
            "Room Chart Report", # New option for room chart
            "Data Processing & Reports",
            "Remuneration Bill Generation",
            "Report Panel"
        ])

        # Conditional rendering based on data availability for core functions
        # Individual functions will now check for data and display warnings.
            
        
        if admin_option == "Get All Students for Date & Shift (Room Wise)":
            st.subheader("List All Students for a Date and Shift (Room Wise)")
            if assigned_seats_df.empty or timetable.empty: # Changed from sitting_plan to assigned_seats_df
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded to use this feature.")
            else:
                list_date_input = st.date_input("Select Date", value=datetime.date.today())
                list_shift_options = ["Morning", "Evening"]
                list_shift = st.selectbox("Select Shift", list_shift_options)
                
                if st.button("Get Student List (Room Wise)"):
                    formatted_student_list_text, error_message, excel_data_for_students_list = get_all_students_for_date_shift_formatted(
                        list_date_input.strftime('%d-%m-%Y'),
                        list_shift,
                        assigned_seats_df, # Pass assigned_seats_df
                        timetable
                    )
                    if formatted_student_list_text:
                        st.success(f"Generated list for {list_date_input.strftime('%d-%m-%Y')} ({list_shift} Shift):")
                        st.text_area("Student List (Text Format)", formatted_student_list_text, height=500)
                        
                        # Download button for TXT
                        file_name_txt = (
                            f"all_students_list_room_wise_{list_date_input.strftime('%Y%m%d')}_"
                            f"{list_shift.lower()}.txt"
                        )
                        st.download_button(
                            label="Download Student List (Room Wise) as TXT",
                            data=formatted_student_list_text,
                            file_name=file_name_txt,
                            mime="text/plain"
                        )

                        # Download button for Excel
                        if excel_data_for_students_list:
                            output = io.BytesIO()
                            workbook = Workbook()
                            sheet = workbook.active
                            sheet.title = "Student List (Room Wise)"

                            for row_data in excel_data_for_students_list:
                                sheet.append(row_data)

                            for col_idx, col_cells in enumerate(sheet.columns):
                                max_length = 0
                                for cell in col_cells:
                                    try:
                                        if cell.value is not None:
                                            cell_value_str = str(cell.value)
                                            current_length = max(len(line) for line in cell_value_str.split('\n'))
                                            if current_length > max_length:
                                                max_length = current_length
                                    except Exception as e:
                                        st.error(f"Error processing cell: {e}")
                                        pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

                            workbook.save(output)
                            processed_data = output.getvalue()

                            file_name_excel = (
                                f"all_students_list_room_wise_{list_date_input.strftime('%Y%m%d')}_"
                                f"{list_shift.lower()}.xlsx"
                            )
                            st.download_button(
                                label="Download Student List (Room Wise) as Excel",
                                data=processed_data,
                                file_name=file_name_excel,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning(f"No students found: {error_message}")

        elif admin_option == "Get All Students for Date & Shift (Roll Number Wise)":
            st.subheader("List All Students for a Date and Shift (Roll Number Wise)")
            if assigned_seats_df.empty or timetable.empty: # Changed from sitting_plan to assigned_seats_df
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded to use this feature.")
            else:
                list_date_input = st.date_input("Select Date", value=datetime.date.today(), key="roll_num_wise_date")
                list_shift_options = ["Morning", "Evening"]
                list_shift = st.selectbox("Select Shift", list_shift_options, key="roll_num_wise_shift")
                
                if st.button("Get Student List (Roll Number Wise)"):
                    formatted_student_list_text, error_message, excel_data_for_students_list = get_all_students_roll_number_wise_formatted(
                        list_date_input.strftime('%d-%m-%Y'),
                        list_shift,
                        assigned_seats_df, # Pass assigned_seats_df
                        timetable
                    )
                    if formatted_student_list_text:
                        st.success(f"Generated list for {list_date_input.strftime('%d-%m-%Y')} ({list_shift} Shift):")
                        st.text_area("Student List (Text Format)", formatted_student_list_text, height=500)
                        
                        # Download button for TXT
                        file_name_txt = (
                            f"all_students_list_roll_wise_{list_date_input.strftime('%Y%m%d')}_"
                            f"{list_shift.lower()}.txt"
                        )
                        st.download_button(
                            label="Download Student List (Roll Number Wise) as TXT",
                            data=formatted_student_list_text,
                            file_name=file_name_txt,
                            mime="text/plain"
                        )

                        # Download button for Excel
                        if excel_data_for_students_list:
                            output = io.BytesIO()
                            workbook = Workbook()
                            sheet = workbook.active
                            sheet.title = "Student List (Roll Wise)"

                            for row_data in excel_data_for_students_list:
                                sheet.append(row_data)

                            for col_idx, col_cells in enumerate(sheet.columns):
                                max_length = 0
                                for cell in col_cells:
                                    try:
                                        if cell.value is not None:
                                            cell_value_str = str(cell.value)
                                            current_length = max(len(line) for line in cell_value_str.split('\n'))
                                            if current_length > max_length:
                                                    max_length = current_length
                                    except Exception as e:
                                        st.error(f"Error processing cell: {e}")
                                        pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

                            workbook.save(output)
                            processed_data = output.getvalue()

                            file_name_excel = (
                                f"all_students_list_roll_wise_{list_date_input.strftime('%Y%m%d')}_"
                                f"{list_shift.lower()}.xlsx"
                            )
                            st.download_button(
                                label="Download Student List (Roll Number Wise) as Excel",
                                data=processed_data,
                                file_name=file_name_excel,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.warning(f"No students found: {error_message}")

        elif admin_option == "Update Timetable Details":
            st.subheader("✏️ Update Timetable Details")
            if timetable.empty:
                st.info("No timetable data loaded. Please upload 'timetable.csv' first using the 'Upload Data Files' section.")
            else:
                st.write("Current Timetable Preview:")
                st.dataframe(timetable)

                st.markdown("---")
                st.write("Select filters to specify which entries to update:")
                
                # Filters for selecting entries to update
                unique_dates_tt = sorted(timetable['Date'].astype(str).unique().tolist())
                unique_shifts_tt = sorted(timetable['Shift'].astype(str).unique().tolist())
                unique_classes_tt = sorted(timetable['Class'].astype(str).unique().tolist())
                unique_paper_codes_tt = sorted(timetable['Paper Code'].astype(str).unique().tolist())
                unique_paper_tt = sorted(timetable['Paper'].astype(str).unique().tolist())
                unique_paper_names_tt = sorted(timetable['Paper Name'].astype(str).unique().tolist())

                filter_date_tt_update = st.selectbox("Filter by Date", ["All"] + unique_dates_tt, key="filter_date_tt_update")
                filter_shift_tt_update = st.selectbox("Filter by Shift", ["All"] + unique_shifts_tt, key="filter_shift_tt_update")
                filter_class_tt_update = st.selectbox("Filter by Class", ["All"] + unique_classes_tt, key="filter_class_tt_update")
                filter_paper_code_tt_update = st.selectbox("Filter by Paper Code", ["All"] + unique_paper_codes_tt, key="filter_paper_code_tt_update")
                filter_paper_tt_update = st.selectbox("Filter by Paper", ["All"] + unique_paper_tt, key="filter_paper_tt_update")
                filter_paper_name_tt_update = st.selectbox("Filter by Paper Name", ["All"] + unique_paper_names_tt, key="filter_paper_name_tt_update")

                st.markdown("---")
                st.write("Entries that will be updated based on your filters:")
                
                temp_filtered_tt = timetable.copy()
                if filter_date_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Date'].astype(str) == filter_date_tt_update]
                if filter_shift_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Shift'].astype(str) == filter_shift_tt_update]
                if filter_class_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Class'].astype(str) == filter_class_tt_update]
                if filter_paper_code_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper Code'].astype(str) == filter_paper_code_tt_update]
                if filter_paper_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper'].astype(str) == filter_paper_tt_update]
                if filter_paper_name_tt_update != "All":
                    temp_filtered_tt = temp_filtered_tt[temp_filtered_tt['Paper Name'].astype(str) == filter_paper_name_tt_update]
                
                if temp_filtered_tt.empty:
                    st.info("No entries match the selected filters. No updates will be applied.")
                else:
                    st.dataframe(temp_filtered_tt)

                st.markdown("---")
                st.write("Enter new values for 'Date', 'Shift', and 'Time' for the filtered entries:")
                
                # Provide default values from the first row of the *filtered* timetable if available, otherwise from the full timetable or current date/time
                default_date_update_input = datetime.date.today()
                if not temp_filtered_tt.empty and 'Date' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Date'].iloc[0]):
                    try:
                        default_date_update_input = datetime.datetime.strptime(str(temp_filtered_tt['Date'].iloc[0]).strip(), '%d-%m-%Y').date()
                    except ValueError:
                        pass
                elif 'Date' in timetable.columns and not timetable['Date'].empty and pd.notna(timetable['Date'].iloc[0]):
                    try:
                        default_date_update_input = datetime.datetime.strptime(str(timetable['Date'].iloc[0]).strip(), '%d-%m-%Y').date()
                    except ValueError:
                        pass


                default_shift_update_input = "Morning"
                if not temp_filtered_tt.empty and 'Shift' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Shift'].iloc[0]):
                    default_shift_update_input = str(temp_filtered_tt['Shift'].iloc[0]).strip()
                elif 'Shift' in timetable.columns and not timetable['Shift'].empty and pd.notna(timetable['Shift'].iloc[0]):
                    default_shift_update_input = str(timetable['Shift'].iloc[0]).strip()


                default_time_update_input = "09:00 AM - 12:00 PM"
                if not temp_filtered_tt.empty and 'Time' in temp_filtered_tt.columns and pd.notna(temp_filtered_tt['Time'].iloc[0]):
                    default_time_update_input = str(temp_filtered_tt['Time'].iloc[0]).strip()
                elif 'Time' in timetable.columns and not timetable['Time'].empty and pd.notna(timetable['Time'].iloc[0]):
                    default_time_update_input = str(timetable['Time'].iloc[0]).strip()


                update_date = st.date_input("New Date", value=default_date_update_input, key="update_tt_date")
                update_shift = st.selectbox("New Shift", ["Morning", "Evening"], index=["Morning", "Evening"].index(default_shift_update_input) if default_shift_update_input in ["Morning", "Evening"] else 0, key="update_tt_shift")
                update_time = st.text_input("New Time (e.g., 09:00 AM - 12:00 PM)", value=default_time_update_input, key="update_tt_time")

                if st.button("Apply Updates and Save Timetable"):
                    if temp_filtered_tt.empty:
                        st.warning("No entries matched your filters, so no updates were applied.")
                    else:
                        timetable_modified = timetable.copy()
                        
                        # Identify indices to update in the original DataFrame
                        indices_to_update = timetable_modified[
                            (timetable_modified['Date'].astype(str) == filter_date_tt_update if filter_date_tt_update != "All" else True) &
                            (timetable_modified['Shift'].astype(str) == filter_shift_tt_update if filter_shift_tt_update != "All" else True) &
                            (timetable_modified['Class'].astype(str) == filter_class_tt_update if filter_class_tt_update != "All" else True) &
                            (timetable_modified['Paper Code'].astype(str) == filter_paper_code_tt_update if filter_paper_code_tt_update != "All" else True) &                           (timetable_modified['Paper'].astype(str) == filter_paper_tt_update if filter_paper_tt_update != "All" else True) &
                            (timetable_modified['Paper Name'].astype(str) == filter_paper_name_tt_update if filter_paper_name_tt_update != "All" else True)
                        ].index

                        # Apply updates only to the identified rows
                        if not indices_to_update.empty:
                            timetable_modified.loc[indices_to_update, 'Date'] = update_date.strftime('%d-%m-%Y')
                            timetable_modified.loc[indices_to_update, 'Shift'] = update_shift
                            timetable_modified.loc[indices_to_update, 'Time'] = update_time

                            success, msg = save_uploaded_file(timetable_modified, TIMETABLE_FILE)
                            if success:
                                st.success(f"Timetable details updated for {len(indices_to_update)} entries and saved successfully.")
                                # Reload data to reflect changes in the app
                                sitting_plan, timetable, assigned_seats_df = load_data() 
                                st.rerun()
                            else:
                                st.error(msg)
                        else:
                            st.warning("No entries matched your filters to apply updates.")

        elif admin_option == "Assign Rooms & Seats to Students": # Replaced with assign_seats_app.py logic
            st.subheader("📘 Room & Seat Assignment Tool")
            st.markdown("""
            This tool helps manage seat assignments for exams, offering real-time status updates,
            capacity warnings, and clear error messages based on your selected seat format.
            """)

            if sitting_plan.empty or timetable.empty:
                st.error(f"Error: `{SITTING_PLAN_FILE}` or `{TIMETABLE_FILE}` not found. Please upload these files to run the assignment tool.")
                st.stop() # Stop execution if critical files are missing

            # --- Session State for consistent UI updates ---
            if 'current_room_status_a_rem' not in st.session_state:
                st.session_state.current_room_status_a_rem = None
            if 'current_room_status_b_rem' not in st.session_state:
                st.session_state.current_room_status_b_rem = None
            if 'current_room_status_total_rem' not in st.session_state:
                st.session_state.current_room_status_total_rem = None

            # --- Input Widgets ---
            st.subheader("Exam Details")
            # Ensure date and shift options are available from timetable
            date_options = sorted(timetable["Date"].dropna().unique())
            shift_options = sorted(timetable["Shift"].dropna().unique())

            if not date_options or not shift_options:
                st.warning("Timetable is empty or missing Date/Shift information. Please upload a complete timetable.")
            else:
                date = st.selectbox("Select Exam Date", date_options, key="assign_date_select")
                shift = st.selectbox("Select Shift", shift_options, key="assign_shift_select")

                # --- NEW: Paper-wise summary chart for the selected session ---
                st.markdown("---")
                st.subheader("Session Student Summary (Assigned vs. Unassigned)")
                session_paper_summary_df = get_session_paper_summary(date, shift, sitting_plan, assigned_seats_df, timetable)
                if not session_paper_summary_df.empty:
                    st.dataframe(session_paper_summary_df)
                else:
                    st.info("No student data found for the selected date and shift.")
                st.markdown("---")
                # --- END NEW ---

                # Filter relevant papers based on selected date and shift
                filtered_papers = timetable[(timetable["Date"] == date) & (timetable["Shift"] == shift)]
                # Ensure paper codes are formatted for display
                paper_options = filtered_papers[["Paper Code", "Paper Name"]].drop_duplicates().values.tolist()
                paper_display = [f"{_format_paper_code(code)} - {name}" for code, name in paper_options]

                selected_paper = st.selectbox("Select Paper Code and Name", paper_display, key="assign_paper_select")

                # Only proceed if a paper is selected
                if selected_paper:
                    paper_code = _format_paper_code(selected_paper.split(" - ")[0]) # Format the extracted code
                    paper_name = selected_paper.split(" - ", 1)[1].strip()

                    st.subheader("Room & Seat Configuration")
                    # Added .strip() to handle potential leading/trailing spaces in room input
                    room = st.text_input("Enter Room Number (e.g., 1, G230)", key="room_input").strip()

                    # Enhanced capacity input
                    col1, col2 = st.columns(2)
                    with col1:
                        total_capacity = st.number_input("Enter Total Room Capacity (for '1 to N' format)", min_value=1, max_value=200, value=60, key="total_capacity_input")
                    with col2:
                        capacity_per_format = st.number_input("Capacity per Format (for 'A/B' formats)", min_value=1, max_value=100, value=30, key="capacity_per_format_input")

                    seat_format = st.radio("Select Seat Assignment Format:", ["1 to N", "1A to NA", "1B to NB"], key="seat_format_radio")

                    # --- Show current room status BEFORE assignment ---
                    if room:
                        # Get all assigned seats for the current room, date, and shift
                        room_assigned_seats_current = assigned_seats_df[
                            (assigned_seats_df["Room Number"] == room) &
                            (assigned_seats_df["Date"] == date) &
                            (assigned_seats_df["Shift"] == shift)
                        ]["Seat Number"].tolist()

                        # Calculate used seats for A, B, and no-suffix formats
                        a_seats_used_current = len([s for s in room_assigned_seats_current if str(s).endswith("A") and s])
                        b_seats_used_current = len([s for s in room_assigned_seats_current if str(s).endswith("B") and s])
                        no_suffix_seats_used_current = len([s for s in room_assigned_seats_current if not str(s).endswith("A") and not str(s).endswith("B")])

                        st.subheader("📊 Current Room Status")
                        if seat_format in ["1A to NA", "1B to NB"]:
                            a_remaining_current = capacity_per_format - a_seats_used_current
                            b_remaining_current = capacity_per_format - b_seats_used_current
                            st.info(f"A-format: **{a_remaining_current}** remaining ({a_seats_used_current}/{capacity_per_format} used)")
                            st.info(f"B-format: **{b_remaining_current}** remaining ({b_seats_used_current}/{capacity_per_format} used)")
                            st.session_state.current_room_status_a_rem = a_remaining_current
                            st.session_state.current_room_status_b_rem = b_remaining_current
                            st.session_state.current_room_status_total_rem = None # Clear total if A/B is selected
                        else: # 1 to N format
                            remaining_current = total_capacity - no_suffix_seats_used_current
                            st.info(f"Total: **{remaining_current}** seats remaining ({no_suffix_seats_used_current}/{total_capacity} used)")
                            st.session_state.current_room_status_total_rem = remaining_current
                            st.session_state.current_room_status_a_rem = None # Clear A/B if total is selected
                            st.session_state.current_room_status_b_rem = None


                    st.markdown("---")

                    # --- Assign Seats Button ---
                    if st.button("✅ Assign Seats", key="assign_button"):
                        if not room:
                            st.error("Please enter a Room Number before assigning seats.")
                            st.stop()

                        # Extract roll numbers for the selected paper from sitting_plan
                        roll_cols = [col for col in sitting_plan.columns if col.lower().startswith("roll number")]
                        # Ensure Paper Code is treated as string for comparison
                        paper_rows = sitting_plan[sitting_plan["Paper Code"].astype(str) == paper_code] # Use formatted paper code
                        all_rolls = paper_rows[roll_cols].values.flatten()
                        all_rolls = [str(r).strip() for r in all_rolls if str(r).strip() and str(r).lower() != 'nan']

                        # Remove previously assigned roll numbers for this paper/date/shift
                        already_assigned_rolls = assigned_seats_df[
                            (assigned_seats_df["Paper Code"].astype(str) == paper_code) & # Use formatted paper code
                            (assigned_seats_df["Date"] == date) &
                            (assigned_seats_df["Shift"] == shift)
                        ]["Roll Number"].astype(str).tolist()

                        unassigned_rolls = [r for r in all_rolls if r not in already_assigned_rolls]

                        if not unassigned_rolls:
                            st.warning("⚠️ All students for this paper are already assigned for this date/shift!")
                            st.stop()

                        # Determine seat format and capacity for the assignment logic
                        suffix = ""
                        format_capacity_for_assignment = 0 # Initialize

                        if seat_format == "1 to N":
                            suffix = ""
                            format_capacity_for_assignment = total_capacity
                        elif seat_format == "1A to NA":
                            suffix = "A"
                            format_capacity_for_assignment = capacity_per_format
                        elif seat_format == "1B to NB":
                            suffix = "B"
                            format_capacity_for_assignment = capacity_per_format

                        # Get a set of all *physically occupied seat keys* for the current room, date, and shift
                        occupied_physical_seat_keys = set(
                            (str(x[0]), str(x[1]), str(x[2]), str(x[3]))
                            for x in assigned_seats_df[
                                (assigned_seats_df["Room Number"] == room) &
                                (assigned_seats_df["Date"] == date) &
                                (assigned_seats_df["Shift"] == shift)
                            ][['Room Number', 'Seat Number', 'Date', 'Shift']].values
                        )

                        # Find truly available seat numbers for the selected format.
                        available_seat_numbers = []
                        for i in range(1, format_capacity_for_assignment + 1):
                            prospective_seat_string = f"{i}{suffix}"
                            prospective_seat_key = (str(room), prospective_seat_string, str(date), str(shift)) # Ensure consistency

                            # A seat is available if its specific key (Room, Seat String, Date, Shift) is NOT already taken
                            if prospective_seat_key not in occupied_physical_seat_keys:
                                available_seat_numbers.append(i)

                        # --- Clear Error Messages & No Automatic Format Switching ---
                        if not available_seat_numbers:
                            st.error(f"❌ ERROR: No seats available in **{seat_format}** format for Room {room}! Please manually change to a different format (e.g., '1A to NA' or '1B to NB') or room.")
                            st.stop() # Stop execution after displaying error

                        # --- Capacity Warnings ---
                        if len(available_seat_numbers) < len(unassigned_rolls):
                            st.warning(f"⚠️ Capacity Warning: Only **{len(available_seat_numbers)}** seats available in **{seat_format}** format, but **{len(unassigned_rolls)}** students need assignment.")
                            st.warning(f"💡 This will assign the first **{len(available_seat_numbers)}** students. Remaining students will need assignment in a different format or room.")

                        # Generate actual seat strings with the selected suffix
                        seats_to_assign_count = min(len(available_seat_numbers), len(unassigned_rolls))
                        assigned_seat_strings = [f"{available_seat_numbers[i]}{suffix}" for i in range(seats_to_assign_count)]

                        # Assign seats to students
                        students_to_assign = unassigned_rolls[:seats_to_assign_count]
                        assigned_rows = []

                        for i, roll in enumerate(students_to_assign):
                            seat_num_str = assigned_seat_strings[i]
                            current_assignment_key = (str(room), seat_num_str, str(date), str(shift)) # Ensure consistency

                            # Check if this specific physical seat key is already taken
                            if current_assignment_key in occupied_physical_seat_keys:
                                st.warning(f"⚠️ Conflict: Seat **{seat_num_str}** in Room **{room}** is already assigned for this date/shift. Skipping assignment for Roll Number **{roll}**.")
                            else:
                                assigned_rows.append({
                                    "Roll Number": roll,
                                    "Paper Code": paper_code, # Keep as string
                                    "Paper Name": paper_name,
                                    "Room Number": room,
                                    "Seat Number": seat_num_str,
                                    "Date": date,
                                    "Shift": shift
                                })
                                # Add this new assignment's physical seat key to our occupied set for this batch
                                occupied_physical_seat_keys.add(current_assignment_key) # Update the set for subsequent assignments in this batch

                        new_assignments_df = pd.DataFrame(assigned_rows)

                        if new_assignments_df.empty:
                            st.warning("No new unique seats could be assigned in this attempt, possibly due to conflicts with existing assignments.")
                            # st.stop() # Removed st.stop() to allow further interaction
                        else:
                            # Merge new assignments with existing ones and save
                            assigned_seats_df = pd.concat([assigned_seats_df, new_assignments_df], ignore_index=True)
                            # Re-add drop_duplicates on Roll Number/Paper Code/Date/Shift to prevent a student
                            # from being assigned the *same paper* multiple times if the button is clicked repeatedly.
                            assigned_seats_df.drop_duplicates(subset=["Roll Number", "Paper Code", "Date", "Shift"], inplace=True)
                            
                            success, msg = save_uploaded_file(assigned_seats_df, ASSIGNED_SEATS_FILE)
                            if success:
                                st.success(f"✅ Successfully assigned **{len(new_assignments_df)}** students to Room **{room}** using **{seat_format}** format.")
                                st.dataframe(new_assignments_df) # Display only the newly assigned students
                            else:
                                st.error(f"Error saving assigned seats: {msg}")

                            # --- Display Updated Room Status AFTER assignment ---
                            st.subheader("📊 Updated Room Status")
                            updated_room_assigned_seats = assigned_seats_df[
                                (assigned_seats_df["Room Number"] == room) &
                                (assigned_seats_df["Date"] == date) &
                                (assigned_seats_df["Shift"] == shift)
                            ]["Seat Number"].tolist()

                            updated_a_seats_used = len([s for s in updated_room_assigned_seats if s.endswith("A")])
                            updated_b_seats_used = len([s for s in updated_room_assigned_seats if s.endswith("B")])
                            updated_no_suffix_seats_used = len([s for s in updated_room_assigned_seats if not s.endswith("A") and not s.endswith("B")])

                            if seat_format in ["1A to NA", "1B to NB"]:
                                updated_a_remaining = capacity_per_format - updated_a_seats_used
                                updated_b_remaining = capacity_per_format - updated_b_seats_used
                                st.info(f"A-format: **{updated_a_remaining}** remaining ({updated_a_seats_used}/{capacity_per_format} used)")
                                st.info(f"B-format: **{updated_b_remaining}** remaining ({updated_b_seats_used}/{capacity_per_format} used)")
                            else: # 1 to N format
                                updated_remaining = total_capacity - updated_no_suffix_seats_used
                                st.info(f"Total: **{updated_remaining}** seats remaining ({updated_no_suffix_seats_used}/{total_capacity} used)")

                            if len(new_assignments_df) < len(unassigned_rolls):
                                remaining_students_after_assignment = len(unassigned_rolls) - len(new_assignments_df)
                                st.warning(f"⚠️ **{remaining_students_after_assignment}** students from this paper still need assignment. Please run assignment again, potentially with a different format or room.")
                            
                            st.rerun() # Rerun to refresh the dataframes and status

                    st.markdown("---")

                    # --- Display all assignments for the selected room/date/shift ---
                    if room:
                        with st.expander(f"📄 View all current assignments for Room {room} on {date} ({shift})"):
                            room_assignments_display = assigned_seats_df[
                                (assigned_seats_df["Room Number"] == room) &
                                (assigned_seats_df["Date"] == date) &
                                (assigned_seats_df["Shift"] == shift)
                            ].copy() # Use .copy() to avoid SettingWithCopyWarning

                            if room_assignments_display.empty:
                                st.info("No assignments yet for this room, date, and shift.")
                            else:
                                # Proper sorting for seat numbers (e.g., 1A, 2A, ..., 10A, 1B, 2B)
                                def sort_seat_number(seat):
                                    if isinstance(seat, str):
                                        if seat.endswith('A'):
                                            return (0, int(seat[:-1])) # Group A seats first
                                        elif seat.endswith('B'):
                                            return (1, int(seat[:-1])) # Group B seats second
                                        elif seat.isdigit():
                                            return (2, int(seat)) # Group 1 to N seats last
                                    return (3, seat) # For any unexpected format, put at the end

                                room_assignments_display['sort_key'] = room_assignments_display['Seat Number'].apply(sort_seat_number)
                                room_assignments_sorted = room_assignments_display.sort_values(by='sort_key').drop('sort_key', axis=1)
                                st.dataframe(room_assignments_sorted, use_container_width=True)

            # --- Reset Button (outside the paper selection block for broader access) ---
            st.markdown("---")
            st.subheader("Maintenance")
            if st.button("🔄 Reset All Assigned Seats (Clear assigned_seats.csv)", key="reset_button"):
                if os.path.exists(ASSIGNED_SEATS_FILE):
                    os.remove(ASSIGNED_SEATS_FILE)
                    st.success("`assigned_seats.csv` has been deleted. All assignments reset.")
                else:
                    st.info("No `assigned_seats.csv` found to reset.")
                st.rerun() # Rerun the app to reflect the changes

        elif admin_option == "Room Occupancy Report": # New Room Occupancy Report section
            display_room_occupancy_report(sitting_plan, assigned_seats_df, timetable)
        
        elif admin_option == "Remuneration Bill Generation":
            st.subheader("💰 Remuneration Bill Generation")
            st.info("Calculate remuneration for exam team members based on their assignments.")

            # Load necessary dataframes
            shift_assignments_df = load_shift_assignments()
            room_invigilator_assignments_df = load_room_invigilator_assignments()
            # Ensure 'Date' column is datetime for sorting and grouping in calculate_remuneration
            # The function itself will convert to datetime, here we just load
            
            # Load assigned_seats_df for Class 3/4 worker calculation
            _, _, assigned_seats_df_for_remuneration = load_data()
            # Load timetable for date parsing in duty dates and for class filtering
            _, timetable_df_for_remuneration, _ = load_data()

            if shift_assignments_df.empty and room_invigilator_assignments_df.empty:
                st.warning("No shift or room invigilator assignments found. Please make assignments first.")
                st.stop()
            
            # Get unique classes from the timetable for multi-selection
            all_classes_in_timetable = sorted(timetable_df_for_remuneration['Class'].dropna().astype(str).str.strip().unique().tolist())
            
            st.markdown("---")
            st.subheader("Select Classes for Bill Generation")
            st.info("Select specific classes to include in the remuneration calculation for shift-based roles. Double shift duty conveyance will still be considered for all shifts worked.")
            selected_classes_for_bill = st.multiselect(
                "Select Classes (leave empty for all classes)",
                options=all_classes_in_timetable,
                default=all_classes_in_timetable # Default to all classes selected
            )

            st.markdown("---")
            st.subheader("Manual Remuneration Rates (per shift/day)")

            # Define default rates and allow user to input
            manual_rates = {
                'senior_center_superintendent_rate': st.number_input("Senior Center Superintendent Rate (Rs./day - no conveyance on exam days)", min_value=0, value=200, key="scs_rate"),
                'center_superintendent_rate': st.number_input("Center Superintendent Rate (Rs.)", min_value=0, value=175, key="cs_rate"),
                'assistant_center_superintendent_rate': st.number_input("Assistant Center Superintendent Rate (Rs.)", min_value=0, value=150, key="acs_rate"),
                'permanent_invigilator_rate': st.number_input("Permanent Invigilator Rate (Rs.)", min_value=0, value=100, key="pi_rate"),
                'assistant_permanent_invigilator_rate': st.number_input("Assistant Permanent Invigilator Rate (Rs.)", min_value=0, value=100, key="api_rate"),
                'invigilator_rate': st.number_input("Invigilator Rate (Rs.)", min_value=0, value=100, key="inv_rate"),
                'conveyance_rate': st.number_input("Conveyance Rate (Evening Shift - both shifts worked) (Rs.)", min_value=0, value=100, key="conveyance_rate"),
                'class_3_worker_rate_per_student': st.number_input("Class 3 Worker Rate (per student) (Rs.)", min_value=0.0, value=4.0, key="c3_rate"),
                'class_4_worker_rate_per_student': st.number_input("Class 4 Worker Rate (per student) (Rs.)", min_value=0.0, value=3.0, key="c4_rate"),
                'holiday_conveyance_allowance_rate': st.number_input("Holiday Conveyance Allowance (Rs.)", min_value=0, value=100, key="holiday_conveyance_rate")
            }

            st.markdown("---")
            st.subheader("Preparation and Closing Day Assignments")
            st.info("Specify preparation and closing days for eligible staff with their specific roles. These days will also be checked for holidays for additional conveyance.")

            all_eligible_members = []
            # Collect all unique names from shift assignments for eligible roles
            for _, row in shift_assignments_df.iterrows():
                for role_col in ['senior_center_superintendent', 'center_superintendent', 'assistant_center_superintendent', 'permanent_invigilator']:
                    if role_col in row and isinstance(row[role_col], list):
                        all_eligible_members.extend(row[role_col])
            all_eligible_members = sorted(list(set(all_eligible_members))) # Get unique names

            # Role options for dropdown
            role_options = [
                'senior_center_superintendent',
                'center_superintendent', 
                'assistant_center_superintendent',
                'permanent_invigilator',
                'assistant_permanent_invigilator',
                'invigilator'
            ]
            
            role_display_names = {
                'senior_center_superintendent': 'Senior Center Superintendent',
                'center_superintendent': 'Center Superintendent',
                'assistant_center_superintendent': 'Assistant Center Superintendent', 
                'permanent_invigilator': 'Permanent Invigilator',
                'assistant_permanent_invigilator': 'Assistant Permanent Invigilator',
                'invigilator': 'Invigilator'
            }

            prep_closing_assignments = {}
            if all_eligible_members:
                for member in all_eligible_members:
                    st.markdown(f"**{member}**")
                    
                    # Role selection for this member
                    selected_role = st.selectbox(
                        f"Select Role for {member}",
                        options=role_options,
                        format_func=lambda x: role_display_names[x],
                        key=f"{member}_role_selection"
                    )
                    
                    prep_days_input = st.text_input(f"Preparation Days for {member} as {role_display_names[selected_role]} (comma-separated DD-MM-YYYY dates)", key=f"{member}_prep_days")
                    closing_days_input = st.text_input(f"Closing Days for {member} as {role_display_names[selected_role]} (comma-separated DD-MM-YYYY dates)", key=f"{member}_closing_days")

                    prep_days_list = [d.strip() for d in prep_days_input.split(',') if d.strip()]
                    closing_days_list = [d.strip() for d in closing_days_input.split(',') if d.strip()]
                    
                    # Basic date format validation
                    valid_prep_days = []
                    for d in prep_days_list:
                        try:
                            datetime.datetime.strptime(d, '%d-%m-%Y')
                            valid_prep_days.append(d)
                        except ValueError:
                            st.warning(f"Invalid date format for {d} in preparation days for {member}. Please use DD-MM-YYYY.")
                    
                    valid_closing_days = []
                    for d in closing_days_list:
                        try:
                            datetime.datetime.strptime(d, '%d-%m-%Y')
                            valid_closing_days.append(d)
                        except ValueError:
                            st.warning(f"Invalid date format for {d} in closing days for {member}. Please use DD-MM-YYYY.")

                    prep_closing_assignments[member] = {
                        'role': selected_role,
                        'prep_days': valid_prep_days,
                        'closing_days': valid_closing_days
                    }
            else:
                st.info("No eligible team members found for preparation/closing day assignments.")

            st.markdown("---")
            st.subheader("Holiday Dates for Conveyance Allowance")
            st.info("Holiday conveyance allowance (Rs. 100) will be given for preparation/closing days that fall on holidays.")
            holiday_dates_input = st.text_input("Enter Holiday Dates (comma-separated DD-MM-YYYY dates)", key="holiday_dates_input")
            holiday_dates = [d.strip() for d in holiday_dates_input.split(',') if d.strip()]
            
            valid_holiday_dates = []
            for d in holiday_dates:
                try:
                    datetime.datetime.strptime(d, '%d-%m-%Y')
                    valid_holiday_dates.append(d)
                except ValueError:
                    st.warning(f"Invalid date format for {d} in holiday dates. Please use DD-MM-YYYY.")
            holiday_dates = valid_holiday_dates

            # Display conveyance rules
            st.markdown("---")
            st.subheader("📋 Conveyance Rules Summary")
            st.info("""
            **Senior Center Superintendent:** Rs. 200/day total (not per shift), no conveyance on exam days, Rs. 100 conveyance only on prep/closing days if they are holidays.
            
            **Other Team Members (excluding Class 3/4 workers):** Rs. 100 conveyance only for evening shifts when worked both morning and evening on the same exam day, Rs. 100 conveyance on prep/closing days if they are holidays.
            
            **Class 3/4 Workers:** No conveyance allowance.
            
            **Note:** Preparation and closing day allowances are given per person per role to avoid duplicates when someone works multiple roles.
            """)

            if st.button("Generate Remuneration Bills"):
                if shift_assignments_df.empty:
                    st.warning("Shift assignments data is required to calculate remuneration.")
                elif assigned_seats_df_for_remuneration.empty:
                    st.warning("Assigned seats data is required to calculate remuneration for Class 3/4 workers.")
                else:
                    with st.spinner("Calculating remuneration..."):
                        df_individual_bills, df_role_summary_matrix, df_class_3_4_final_bills = calculate_remuneration(
                            shift_assignments_df,
                            room_invigilator_assignments_df,
                            timetable_df_for_remuneration, # Pass timetable for date parsing and class filtering
                            assigned_seats_df_for_remuneration,
                            manual_rates,
                            prep_closing_assignments,
                            holiday_dates,
                            selected_classes_for_bill # Pass selected classes
                        )

                        st.markdown("### Individual Remuneration Bills")
                        if not df_individual_bills.empty:
                            df_individual_bills_with_total = add_total_row(df_individual_bills)
                            st.dataframe(df_individual_bills_with_total, use_container_width=True)
                        else:
                            st.info("No individual bills generated.")

                        st.markdown("### Role-wise Summary Matrix")
                        if not df_role_summary_matrix.empty:
                            df_role_summary_matrix_with_total = add_total_row(df_role_summary_matrix)
                            st.dataframe(df_role_summary_matrix_with_total, use_container_width=True)
                        else:
                            st.info("No role-wise summary generated.")

                        st.markdown("### Class 3 & Class 4 Worker Bills")
                        if not df_class_3_4_final_bills.empty:
                            df_class_3_4_final_bills_with_total = add_total_row(df_class_3_4_final_bills)
                            st.dataframe(df_class_3_4_final_bills_with_total, use_container_width=True)
                        else:
                            st.info("No Class 3 & 4 worker bills generated.")
                        
                        # Save and Download Bills
                        if not df_individual_bills.empty or not df_role_summary_matrix.empty or not df_class_3_4_final_bills.empty:
                            excel_file_buffer, excel_filename = save_bills_to_excel(
                                df_individual_bills_with_total, 
                                df_role_summary_matrix_with_total, 
                                df_class_3_4_final_bills_with_total
                            )
                            st.download_button(
                                label="Download All Remuneration Bills as Excel",
                                data=excel_file_buffer,
                                file_name=excel_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                            st.success(f"Remuneration bills generated and ready for download as '{excel_filename}'.")
                        else:
                            st.warning("No bills were generated to save.")


        elif admin_option == "Room Chart Report": # New Room Chart Report section
            st.subheader("📄 Room Chart Report")
            st.info("Generate a detailed room chart showing student seating arrangements for a specific exam session.")

            if sitting_plan.empty or timetable.empty or assigned_seats_df.empty:
                st.warning("Please upload 'sitting_plan.csv', 'timetable.csv', and ensure seats are assigned via 'Assign Rooms & Seats to Students' (Admin Panel) to generate this report.")
                st.stop() # Use st.stop() to halt execution if critical data is missing
            
            # Date and Shift filters for the room chart
            chart_date_options = sorted(timetable["Date"].dropna().unique())
            chart_shift_options = sorted(timetable["Shift"].dropna().unique())

            if not chart_date_options or not chart_shift_options:
                st.info("No exam dates or shifts found in the timetable to generate a room chart.")
                st.stop() # Use st.stop() to halt execution if no options

            selected_chart_date = st.selectbox("Select Date", chart_date_options, key="room_chart_date")
            selected_chart_shift = st.selectbox("Select Shift", chart_shift_options, key="room_chart_shift")

            if st.button("Generate Room Chart"):
                with st.spinner("Generating room chart..."):
                    # The generate_room_chart_report function now returns a string message if there's an error
                    room_chart_output = generate_room_chart_report(selected_chart_date, selected_chart_shift, sitting_plan, assigned_seats_df, timetable)
                    
                    # Check if the output is an error message (string) or the actual chart data
                    if room_chart_output and "Error:" in room_chart_output:
                        st.error(room_chart_output) # Display the error message
                    elif room_chart_output:
                        st.text_area("Generated Room Chart", room_chart_output, height=600)
                        
                        # Download button
                        file_name = f"room_chart_{selected_chart_date}_{selected_chart_shift}.csv"
                        st.download_button(
                            label="Download Room Chart as CSV",
                            data=room_chart_output.encode('utf-8'),
                            file_name=file_name,
                            mime="text/csv",
                        )
                    else:
                        st.warning("Could not generate room chart. Please ensure data is complete and assignments are made.")


        elif admin_option == "Data Processing & Reports":
            st.subheader("⚙️ Data Processing & Report Generation")

            st.markdown("---")
            st.subheader("Upload PDFs for Sitting Plan (pdf_folder.zip)")
            st.info(f"Upload a ZIP file containing folders of PDFs (e.g., 'pdf_folder/Zoology'). Each folder name will be used as the 'Paper' name. This will generate/update '{SITTING_PLAN_FILE}' and an initial '{TIMETABLE_FILE}'.")
            uploaded_sitting_plan_zip = st.file_uploader("Upload Sitting Plan PDFs (ZIP)", type=["zip"], key="upload_sitting_plan_zip")
            if uploaded_sitting_plan_zip:
                with st.spinner("Processing sitting plan PDFs and generating initial timetable... This may take a while."):
                    success, message = process_sitting_plan_pdfs(uploaded_sitting_plan_zip, SITTING_PLAN_FILE, TIMETABLE_FILE)
                    if success:
                        st.success(message)
                        # Reload data after processing
                        sitting_plan, timetable, assigned_seats_df = load_data()
                    else:
                        st.error(message)

            st.markdown("---")
            st.subheader("Upload Attestation PDFs (rasa_pdf.zip)")
            st.info(f"Upload a ZIP file containing attestation PDFs. These will be parsed to create a combined attestation data CSV ('{ATTESTATION_DATA_FILE}') and then automatically generate college statistics ('{COLLEGE_STATISTICS_FILE}').")
            uploaded_attestation_zip = st.file_uploader("Upload Attestation PDFs (ZIP)", type=["zip"], key="upload_attestation_zip")
            if uploaded_attestation_zip:
                with st.spinner("Processing attestation PDFs and generating college statistics... This may take a while."):
                    success, message = process_attestation_pdfs(uploaded_attestation_zip, ATTESTATION_DATA_FILE)
                    if success:
                        st.success(message)
                        # Automatically generate college statistics after attestation PDFs are processed
                        st.info("Generating college statistics...")
                        stats_success, stats_message = generate_college_statistics(ATTESTATION_DATA_FILE, COLLEGE_STATISTICS_FILE)
                        if stats_success:
                            st.success(stats_message)
                            if os.path.exists(COLLEGE_STATISTICS_FILE):
                                with open(COLLEGE_STATISTICS_FILE, "rb") as f:
                                    st.download_button(
                                        label="Download College Statistics CSV",
                                        data=f,
                                        file_name=COLLEGE_STATISTICS_FILE,
                                        mime="text/csv",
                                        key="download_college_stats_auto" # Unique key added
                                    )
                        else:
                            st.error(stats_message)
                    else:
                        st.error(message)

            st.markdown("---")
            st.subheader("Generate College Statistics (Manual Trigger)")
            st.info(f"This will generate college-wise statistics from '{ATTESTATION_DATA_FILE}' and save it to '{COLLEGE_STATISTICS_FILE}'. Only use if attestation data is already processed.")
            if st.button("Generate College Statistics (Manual)"):
                success, message = generate_college_statistics(ATTESTATION_DATA_FILE, COLLEGE_STATISTICS_FILE)
                if success:
                    st.success(message)
                    if os.path.exists(COLLEGE_STATISTICS_FILE):
                        with open(COLLEGE_STATISTICS_FILE, "rb") as f:
                            st.download_button(
                                label="Download College Statistics CSV",
                                data=f,
                                file_name=COLLEGE_STATISTICS_FILE,
                                mime="text/csv",
                                key="download_college_stats_manual" # Unique key added
                            )
                else:
                    st.error(message)

        elif admin_option == "Report Panel":
            display_report_panel()

    else:
        st.warning("Enter valid admin credentials.")

elif menu == "Centre Superintendent Panel":
    st.subheader("🔐 Centre Superintendent Login")
    if cs_login():
        st.success("Login successful!")

        # Load data for CS panel
        sitting_plan, timetable, assigned_seats_df = load_data() # Load assigned_seats_df here
        
        cs_panel_option = st.radio("Select CS Task:", ["Report Exam Session", "Manage Exam Team & Shift Assignments", "View Full Timetable", "Room Chart Report"]) # Added Room Chart Report

        if cs_panel_option == "Manage Exam Team & Shift Assignments":
            st.subheader("👥 Manage Exam Team Members")
            
            current_members = load_exam_team_members()
            new_member_name = st.text_input("Add New Team Member Name")
            if st.button("Add Member"):
                if new_member_name and new_member_name not in current_members:
                    current_members.append(new_member_name)
                    success, msg = save_exam_team_members(current_members)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
                elif new_member_name:
                    st.warning("Member already exists.")
                else:
                    st.warning("Please enter a name.")

            if current_members:
                st.write("Current Team Members:")
                st.write(current_members)
                
                member_to_remove = st.selectbox("Select Member to Remove", [""] + current_members)
                if st.button("Remove Selected Member"):
                    if member_to_remove:
                        current_members.remove(member_to_remove)
                        success, msg = save_exam_team_members(current_members)
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Please select a member to remove.")
            else:
                st.info("No team members added yet.")

            st.markdown("---")
            st.subheader("🗓️ Assign Roles for Exam Shift")

            assignment_date = st.date_input("Select Date for Assignment", value=datetime.date.today(), key="assignment_date")
            assignment_shift = st.selectbox("Select Shift for Assignment", ["Morning", "Evening"], key="assignment_shift")

            all_team_members = load_exam_team_members()
            if not all_team_members:
                st.warning("Please add exam team members first in the 'Manage Exam Team Members' section.")
            else:
                current_assignments_df = load_shift_assignments()
                current_assignment_for_shift = current_assignments_df[
                    (current_assignments_df['date'] == assignment_date.strftime('%d-%m-%Y')) &
                    (current_assignments_df['shift'] == assignment_shift)
                ]
                
                loaded_senior_cs = []
                loaded_cs = []
                loaded_assist_cs = []
                loaded_perm_inv = []
                loaded_assist_perm_inv = []
                loaded_class_3 = []
                loaded_class_4 = []


                if not current_assignment_for_shift.empty:
                    loaded_senior_cs = current_assignment_for_shift.iloc[0].get('senior_center_superintendent', [])
                    loaded_cs = current_assignment_for_shift.iloc[0].get('center_superintendent', [])
                    loaded_assist_cs = current_assignment_for_shift.iloc[0].get('assistant_center_superintendent', [])
                    loaded_perm_inv = current_assignment_for_shift.iloc[0].get('permanent_invigilator', [])
                    loaded_assist_perm_inv = current_assignment_for_shift.iloc[0].get('assistant_permanent_invigilator', [])
                    loaded_class_3 = current_assignment_for_shift.iloc[0].get('class_3_worker', [])
                    loaded_class_4 = current_assignment_for_shift.iloc[0].get('class_4_worker', [])


                selected_senior_cs = st.multiselect("Senior Center Superintendent (Max 1)", all_team_members, default=loaded_senior_cs, max_selections=1)
                selected_cs = st.multiselect("Center Superintendent (Max 1)", all_team_members, default=loaded_cs, max_selections=1)
                selected_assist_cs = st.multiselect("Assistant Center Superintendent (Max 3)", all_team_members, default=loaded_assist_cs, max_selections=3)
                selected_perm_inv = st.multiselect("Permanent Invigilator (Max 1)", all_team_members, default=loaded_perm_inv, max_selections=1)
                selected_assist_perm_inv = st.multiselect("Assistant Permanent Invigilator (Max 5)", all_team_members, default=loaded_assist_perm_inv, max_selections=5)
                selected_class_3 = st.multiselect("Class 3 Worker (Max 10)", all_team_members, default=loaded_class_3, max_selections=10)
                selected_class_4 = st.multiselect("Class 4 Worker (Max 10)", all_team_members, default=loaded_class_4, max_selections=10)


                if st.button("Save Shift Assignments"):
                    all_selected_members = (
                        selected_senior_cs + selected_cs + selected_assist_cs +
                        selected_perm_inv + selected_assist_perm_inv + selected_class_3 + selected_class_4
                    )
                    if len(all_selected_members) != len(set(all_selected_members)):
                        st.error("Error: A team member cannot be assigned to multiple roles for the same shift.")
                    else:
                        assignments_to_save = {
                            'senior_center_superintendent': selected_senior_cs,
                            'center_superintendent': selected_cs,
                            'assistant_center_superintendent': selected_assist_cs,
                            'permanent_invigilator': selected_perm_inv,
                            'assistant_permanent_invigilator': selected_assist_perm_inv,
                            'class_3_worker': selected_class_3,
                            'class_4_worker': selected_class_4
                        }
                        success, msg = save_shift_assignment(assignment_date.strftime('%d-%m-%Y'), assignment_shift, assignments_to_save)
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                
                st.markdown("---")
                st.subheader("Current Shift Assignments")
                display_assignments_df = load_shift_assignments()
                if not display_assignments_df.empty:
                    st.dataframe(display_assignments_df)
                else:
                    st.info("No shift assignments saved yet.")

            st.markdown("---")

            st.subheader("Assign Invigilators to Rooms")
            if assigned_seats_df.empty: # Check assigned_seats_df instead of sitting_plan/timetable for rooms
                st.info("Please assign seats to students via the Admin Panel's 'Assign Rooms & Seats to Students' section first to see available rooms for invigilator assignment.")
            else:
                room_inv_date = st.date_input("Select Date for Room Invigilators", value=datetime.date.today(), key="room_inv_date")
                room_inv_shift = st.selectbox("Select Shift for Room Invigilators", ["Morning", "Evening"], key="room_inv_shift")
                
                # MODIFIED: Get unique rooms for the selected date and shift from assigned_seats_df
                relevant_rooms_assigned = assigned_seats_df[
                    (assigned_seats_df["Date"].astype(str).str.strip() == room_inv_date.strftime('%d-%m-%Y')) &
                    (assigned_seats_df["Shift"].astype(str).str.strip().str.lower() == room_inv_shift.lower())
                ]
                
                unique_relevant_rooms = sorted(list(relevant_rooms_assigned['Room Number'].dropna().astype(str).str.strip().unique()))

                selected_room_for_inv = st.selectbox("Select Room to Assign Invigilators", [""] + unique_relevant_rooms, key="selected_room_for_inv")

                if selected_room_for_inv:
                    current_room_inv_df = load_room_invigilator_assignments()
                    loaded_invigilators = []
                    
                    filtered_inv_for_room = current_room_inv_df[
                        (current_room_inv_df['date'] == room_inv_date.strftime('%d-%m-%Y')) &
                        (current_room_inv_df['shift'] == room_inv_shift) &
                        (current_room_inv_df['room_num'] == selected_room_for_inv)
                    ]
                    
                    if not filtered_inv_for_room.empty:
                        loaded_invigilators = filtered_inv_for_room.iloc[0].get('invigilators', [])
                    
                    invigilators_for_room = st.multiselect(
                        "Invigilators for this Room",
                        options=all_team_members, # Use the same team members list
                        default=loaded_invigilators,
                        key="invigilators_for_room_multiselect"
                    )

                    if st.button("Save Room Invigilators"):
                        success, msg = save_room_invigilator_assignment(
                            room_inv_date.strftime('%d-%m-%Y'),
                            room_inv_shift,
                            selected_room_for_inv,
                            invigilators_for_room
                        )
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                else:
                    st.info("Select a date, shift, and room to assign invigilators.")
                
                st.markdown("---")
                st.subheader("Current Room Invigilator Assignments")
                display_room_inv_df = load_room_invigilator_assignments()
                if not display_room_inv_df.empty:
                    st.dataframe(display_room_inv_df)
                else:
                    st.info("No room invigilator assignments saved yet.")


        elif cs_panel_option == "Report Exam Session":
            st.subheader("📝 Report Exam Session")
            if assigned_seats_df.empty or timetable.empty: # Check assigned_seats_df instead of sitting_plan
                st.info("Please ensure seats are assigned and 'timetable.csv' is uploaded via the Admin Panel to report exam sessions.")
            else:
                # Date and Shift selection
                report_date = st.date_input("Select Date", value=datetime.date.today(), key="cs_report_date")
                report_shift = st.selectbox("Select Shift", ["Morning", "Evening"], key="cs_report_shift")

                # Filter assigned_seats_df for selected date and shift to get available exam sessions
                available_sessions_assigned = assigned_seats_df[
                    (assigned_seats_df["Date"].astype(str).str.strip() == report_date.strftime('%d-%m-%Y')) &
                    (assigned_seats_df["Shift"].astype(str).str.strip().str.lower() == report_shift.lower())
                ].copy()

                if available_sessions_assigned.empty:
                    st.warning("No assigned seats found for the selected date and shift. Please assign seats via the Admin Panel first.")
                else:
                    # Create a unique identifier for each exam session (Room - Paper Code (Paper Name))
                    # Ensure Paper Code and Paper Name are strings before combining
                    available_sessions_assigned['Paper Code'] = available_sessions_assigned['Paper Code'].astype(str)
                    available_sessions_assigned['Paper Name'] = available_sessions_assigned['Paper Name'].astype(str)

                    available_sessions_assigned['exam_session_id'] = \
                        available_sessions_assigned['Room Number'].astype(str).str.strip() + " - " + \
                        available_sessions_assigned['Paper Code'].apply(_format_paper_code) + " (" + \
                        available_sessions_assigned['Paper Name'].str.strip() + ")"
                    
                    unique_exam_sessions = available_sessions_assigned[['Room Number', 'Paper Code', 'Paper Name', 'exam_session_id']].drop_duplicates().sort_values(by='exam_session_id')
                    
                    if unique_exam_sessions.empty:
                        st.warning("No unique exam sessions found for the selected date and shift in assigned seats.")
                    else:
                        selected_exam_session_option = st.selectbox(
                            "Select Exam Session (Room - Paper Code (Paper Name))",
                            [""] + unique_exam_sessions['exam_session_id'].tolist(),
                            key="cs_exam_session_select"
                        )

                        if selected_exam_session_option:
                            # Extract room_number, paper_code, paper_name from the selected option
                            selected_room_num = selected_exam_session_option.split(" - ")[0].strip()
                            selected_paper_code_with_name = selected_exam_session_option.split(" - ", 1)[1].strip()
                            selected_paper_code = _format_paper_code(selected_paper_code_with_name.split(" (")[0]) # Format the extracted code
                            selected_paper_name = selected_paper_code_with_name.split(" (")[1].replace(")", "").strip()

                            # Find the corresponding class for the selected session from timetable
                            # This assumes a paper code/name maps to a consistent class in the timetable
                            matching_class_info = timetable[
                                (timetable['Paper Code'].astype(str).str.strip() == selected_paper_code) & # Use formatted paper code
                                (timetable['Paper Name'].astype(str).str.strip() == selected_paper_name)
                            ]
                            selected_class = ""
                            if not matching_class_info.empty:
                                selected_class = str(matching_class_info.iloc[0]['Class']).strip()

                            # Create a unique key for CSV row ID
                            report_key = f"{report_date.strftime('%Y%m%d')}_{report_shift.lower()}_{selected_room_num}_{selected_paper_code}"

                            # Load existing report from CSV
                            loaded_success, loaded_report = load_single_cs_report_csv(report_key)
                            if loaded_success:
                                st.info("Existing report loaded.")
                            else:
                                st.info("No existing report found for this session. Starting new.")
                                loaded_report = {} # Ensure it's an empty dict if not found

                            # MODIFIED: Get all *assigned* roll numbers for this specific session from assigned_seats_df
                            expected_students_for_session = assigned_seats_df[
                                (assigned_seats_df['Room Number'].astype(str).str.strip() == selected_room_num) &
                                (assigned_seats_df['Paper Code'].astype(str).str.strip() == selected_paper_code) & # Use formatted paper code
                                (assigned_seats_df['Paper Name'].astype(str).str.strip() == selected_paper_name) &
                                (assigned_seats_df['Date'].astype(str).str.strip() == report_date.strftime('%d-%m-%Y')) &
                                (assigned_seats_df['Shift'].astype(str).str.strip().str.lower() == report_shift.lower())
                            ]['Roll Number'].astype(str).tolist()
                            
                            expected_students_for_session = sorted(list(set(expected_students_for_session))) # Remove duplicates and sort

                            st.write(f"**Reporting for:** Room {selected_room_num}, Paper: {selected_paper_name} ({selected_paper_code})")

                            # Multiselect for Absent Roll Numbers
                            absent_roll_numbers_selected = st.multiselect(
                                "Absent Roll Numbers", 
                                options=expected_students_for_session, 
                                default=loaded_report.get('absent_roll_numbers', []),
                                key="absent_roll_numbers_multiselect"
                            )

                            # Multiselect for UFM Roll Numbers
                            ufm_roll_numbers_selected = st.multiselect(
                                "UFM (Unfair Means) Roll Numbers", 
                                options=expected_students_for_session, 
                                default=loaded_report.get('ufm_roll_numbers', []),
                                key="ufm_roll_numbers_multiselect"
                            )

                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button("Save Report", key="save_cs_report"):
                                    # --- Validation Logic ---
                                    expected_set = set(expected_students_for_session)
                                    absent_set = set(absent_roll_numbers_selected)
                                    ufm_set = set(ufm_roll_numbers_selected)

                                    validation_errors = []

                                    # 1. All reported absent students must be in the expected list
                                    if not absent_set.issubset(expected_set):
                                        invalid_absent = list(absent_set.difference(expected_set))
                                        validation_errors.append(f"Error: Absent roll numbers {invalid_absent} are not in the expected student list for this session.")

                                    # 2. All reported UFM students must be in the expected list
                                    if not ufm_set.issubset(expected_set):
                                        invalid_ufm = list(ufm_set.difference(expected_set))
                                        validation_errors.append(f"Error: UFM roll numbers {invalid_ufm} are not in the expected student list for this session.")

                                    # 3. No student can be both absent and UFM
                                    if not absent_set.isdisjoint(ufm_set):
                                        overlap = list(absent_set.intersection(ufm_set))
                                        validation_errors.append(f"Error: Roll numbers {overlap} are marked as both Absent and UFM. A student cannot be both.")
                                    
                                    if validation_errors:
                                        for err in validation_errors:
                                            st.error(err)
                                    else:
                                        report_data = {
                                            'report_key': report_key, # Add report_key to data
                                            'date': report_date.strftime('%d-%m-%Y'),
                                            'shift': report_shift,
                                            'room_num': selected_room_num,
                                            'paper_code': selected_paper_code,
                                            'paper_name': selected_paper_name,
                                            'class': selected_class, # Added 'class' here
                                            'absent_roll_numbers': absent_roll_numbers_selected, # Store as list
                                            'ufm_roll_numbers': ufm_roll_numbers_selected # Store as list
                                        }
                                        success, message = save_cs_report_csv(report_key, report_data)
                                        if success:
                                            st.success(message)
                                        else:
                                            st.error(message)
                                        st.rerun() # Rerun to refresh the UI with saved data

                                st.markdown("---")
                                st.subheader("All Saved Reports (for debugging/review)")
                                
                                # Fetch all reports for the current CS user from CSV
                                all_reports_df_display = load_cs_reports_csv()
                                room_invigilators_df_display = load_room_invigilator_assignments()

                                if not all_reports_df_display.empty:
                                    # Merge with room invigilators for display
                                    if not room_invigilators_df_display.empty:
                                        all_reports_df_display = pd.merge(
                                            all_reports_df_display,
                                            room_invigilators_df_display[['date', 'shift', 'room_num', 'invigilators']],
                                            on=['date', 'shift', 'room_num'],
                                            how='left',
                                            suffixes=('', '_room_inv_display')
                                        )
                                        all_reports_df_display['invigilators'] = all_reports_df_display['invigilators'].apply(lambda x: x if isinstance(x, list) else [])
                                    else:
                                        all_reports_df_display['invigilators'] = [[]] * len(all_reports_df_display)

                                    # Reorder columns for better readability
                                    display_cols = [
                                        "date", "shift", "room_num", "paper_code", "paper_name", "class", 
                                        "invigilators", "absent_roll_numbers", "ufm_roll_numbers", "report_key"
                                    ]
                                    # Map internal keys to display keys
                                    df_all_reports_display = all_reports_df_display.rename(columns={
                                        'date': 'Date', 'shift': 'Shift', 'room_num': 'Room',
                                        'paper_code': 'Paper Code', 'paper_name': 'Paper Name', 'class': 'Class', 
                                        'invigilators': 'Invigilators',
                                        'absent_roll_numbers': 'Absent Roll Numbers',
                                        'ufm_roll_numbers': 'UFM Roll Numbers',
                                        'report_key': 'Report Key'
                                    })
                                    
                                    # Ensure all display_cols exist, fill missing with empty string
                                    for col in display_cols:
                                        if col not in df_all_reports_display.columns:
                                            df_all_reports_display[col] = ""
                                    
                                    st.dataframe(df_all_reports_display[
                                        ['Date', 'Shift', 'Room', 'Paper Code', 'Paper Name', 'Class', 
                                         'Invigilators', 'Absent Roll Numbers', 'UFM Roll Numbers', 'Report Key']
                                    ])
                                else:
                                    st.info("No reports saved yet.")

        elif cs_panel_option == "View Full Timetable": # New section for CS timetable view
            st.subheader("Full Examination Timetable")
            if timetable.empty:
                st.warning("Timetable data is missing. Please upload it via the Admin Panel.")
            else:
                st.dataframe(timetable)

        elif cs_panel_option == "Room Chart Report": # New Room Chart Report section for CS
            st.subheader("📄 Room Chart Report")
            st.info("Generate a detailed room chart showing student seating arrangements for a specific exam session.")

            if sitting_plan.empty or timetable.empty or assigned_seats_df.empty:
                st.warning("Please upload 'sitting_plan.csv', 'timetable.csv', and ensure seats are assigned via 'Assign Rooms & Seats to Students' (Admin Panel) to generate this report.")
                st.stop() # Use st.stop() to halt execution if critical data is missing
            
            # Date and Shift filters for the room chart
            chart_date_options = sorted(timetable["Date"].dropna().unique())
            chart_shift_options = sorted(timetable["Shift"].dropna().unique())

            if not chart_date_options or not chart_shift_options:
                st.info("No exam dates or shifts found in the timetable to generate a room chart.")
                st.stop() # Use st.stop() to halt execution if no options

            selected_chart_date = st.selectbox("Select Date", chart_date_options, key="cs_room_chart_date")
            selected_chart_shift = st.selectbox("Select Shift", chart_shift_options, key="cs_room_chart_shift")

            if st.button("Generate Room Chart"):
                with st.spinner("Generating room chart..."):
                    # The generate_room_chart_report function now returns a string message if there's an error
                    room_chart_output = generate_room_chart_report(selected_chart_date, selected_chart_shift, sitting_plan, assigned_seats_df, timetable)
                    
                    # Check if the output is an error message (string) or the actual chart data
                    if room_chart_output and "Error:" in room_chart_output:
                        st.error(room_chart_output) # Display the error message
                    elif room_chart_output:
                        st.text_area("Generated Room Chart", room_chart_output, height=600)
                        
                        # Download button
                        file_name = f"room_chart_{selected_chart_date}_{selected_chart_shift}.csv"
                        st.download_button(
                            label="Download Room Chart as CSV",
                            data=room_chart_output.encode('utf-8'),
                            file_name=file_name,
                            mime="text/csv",
                        )
                    else:
                        st.warning("Could not generate room chart. Please ensure data is complete and assignments are made.")

    else:
        st.warning("Enter valid Centre Superintendent credentials.")
def delete_file_app():
    """
    Streamlit application to delete a specified file (abc.csv).
    """
    st.title("File Deletion App")
    st.write("This app allows you to delete the 'abc.csv' file from the current directory.")

    file_to_delete = "timetable.csv"

    # Check if the file exists
    if os.path.exists(file_to_delete):
        st.info(f"The file '{file_to_delete}' currently exists.")
        if st.button(f"Delete {file_to_delete}"):
            try:
                os.remove(file_to_delete)
                st.success(f"Successfully deleted '{file_to_delete}'.")
            except OSError as e:
                st.error(f"Error: Could not delete '{file_to_delete}'. Reason: {e}")
            # Re-run the app to update the file existence status
            st.experimental_rerun()
    else:
        st.warning(f"The file '{file_to_delete}' does not exist in the current directory.")
        st.info("You might need to create it first for the delete button to appear.")

if __name__ == "__main__":
    delete_file_app()
